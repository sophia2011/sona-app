# =============================================================================
#  CONSTRUCTOR PRO  ·  v4  ·  Control de Obras + CRM + Compras
#  SONA CONSTRUCTORA DEL MAYAB
#
#  CAMBIOS v4:
#    - La app inicia VACÍA (como recién adquirida): sin datos de ejemplo.
#      Tú das de alta tus clientes, obras, contratistas, etc.
#    - Nuevo apartado COMPRAS: registro de compras/gastos diarios con
#      categoría (Materiales, Herramienta/equipo, Comidas, Traslados,
#      Hospedaje, Viáticos), quién compra, fecha, y adjuntar COMPROBANTE
#      y FACTURA. La compra se carga a la OBRA ACTIVA.
#    - Botón para BORRAR TODOS LOS DATOS (en Editar / Borrar, Admin).
#
#  -------------------------------------------------------------------------
#  INSTALACIÓN:   python -m pip install streamlit pandas plotly
#  (opcional Excel): python -m pip install openpyxl
#  PARA CORRER:   python -m streamlit run app.py
#  -------------------------------------------------------------------------
# =============================================================================

import os
import shutil
import sqlite3
import hashlib
import urllib.parse
from datetime import date, datetime, timedelta, timezone

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

try:
    from fpdf import FPDF
    FPDF_OK = True
except Exception:
    FPDF_OK = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSHEETS_OK = True
except Exception:
    GSHEETS_OK = False

try:
    import openpyxl  # noqa: F401
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

GCP_JSON = None  # se define tras BASE_DIR

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "cob_data.db")
COMPROB_DIR = os.path.join(BASE_DIR, "comprobantes")  # archivos adjuntos
EMPRESA = "SONA CONSTRUCTORES DEL MAYAB"
LOGO_PATH = os.path.join(BASE_DIR, "logo_sona.png")  # logo para los reportes (opcional)
GCP_JSON = os.path.join(BASE_DIR, ".gcp_service_account.json")  # credenciales Google (opcional)
TABLAS_SYNC = ["clientes", "obras", "etapas", "compras", "requisiciones", "destajos",
               "bitacora", "contratistas", "presupuesto", "proveedores", "usuarios", "abonos"]


def pesos(valor) -> str:
    """Formato de moneda mexicano: $ 1'700,000.00 (apóstrofo para millones)."""
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return "$ 0.00"
    signo = "-" if v < 0 else ""
    izq, der = f"{abs(v):,.2f}".split(".")
    grupos = izq.split(",")
    if len(grupos) >= 3:
        izq = ",".join(grupos[:-2]) + "'" + grupos[-2] + "," + grupos[-1]
    return f"{signo}$ {izq}.{der}"


# =============================================================================
# 0) TEMA CLARO AUTOMÁTICO
# =============================================================================
def asegurar_tema_claro() -> None:
    cfg_dir = os.path.join(BASE_DIR, ".streamlit")
    cfg_file = os.path.join(cfg_dir, "config.toml")
    if not os.path.exists(cfg_file):
        try:
            os.makedirs(cfg_dir, exist_ok=True)
            with open(cfg_file, "w", encoding="utf-8") as f:
                f.write('[theme]\nbase = "light"\nprimaryColor = "#2F6F6A"\n'
                        'backgroundColor = "#F4F2EE"\nsecondaryBackgroundColor = "#FFFFFF"\n'
                        'textColor = "#2A2A28"\nfont = "sans serif"\n')
        except Exception:
            pass


asegurar_tema_claro()

# =============================================================================
# 1) CONFIGURACIÓN GENERAL
# =============================================================================
st.set_page_config(page_title="Constructor PRO · CRM y Obras",
                   page_icon="🏗️", layout="wide", initial_sidebar_state="expanded")

HOY = date.today()


def ahora_mx() -> datetime:
    """Fecha y hora actual en horario del centro de México (UTC-6)."""
    return datetime.now(timezone.utc) - timedelta(hours=6)
COLOR_PRIMARIO = "#2F6F6A"
COLOR_ACENTO = "#C9842B"
COLOR_OK = "#4F8A5B"
COLOR_ALERTA = "#B4554C"

CATEGORIAS_COMPRA = ["Materiales", "Herramienta y/o equipo", "Gasolina", "Comidas",
                     "Traslados", "Hospedaje", "Viáticos"]
ESTATUS_REQ = ["Solicitada", "Aprobada", "Comprada", "Entregada"]
PRIORIDAD_REQ = ["Normal", "Urgente", "En espera"]
TIPO_CLIENTE = ["Prospecto", "Activo", "Cerrado"]
METODOS_PAGO = ["Efectivo", "Transferencia", "Tarjeta de débito", "Tarjeta de crédito"]


# =============================================================================
# 2) ESTILOS (CSS)
# =============================================================================
def inyectar_estilos() -> None:
    st.markdown("""
    <style>
        html, body, [class*="css"] { font-family: 'Segoe UI', system-ui, sans-serif; }
        .stApp { background-color: #F4F2EE; }
        .stApp, .stApp p, .stApp li, .stApp label,
        .stApp h1, .stApp h2, .stApp h3, .stApp h4,
        [data-testid="stMarkdownContainer"] *, [data-testid="stWidgetLabel"] *,
        [data-testid="stHeadingWithActionElements"] * { color: #2A2A28 !important; }
        [data-testid="stSidebar"] { background: #FFFFFF; }
        [data-testid="stSidebar"] * { color: #2A2A28 !important; }
        .stApp input, .stApp textarea, .stApp [data-baseweb="select"] > div {
            background-color: #FFFFFF !important; color: #2A2A28 !important; }
        div[data-testid="stMetric"] {
            background:#FFFFFF; border:1px solid #E6E2DA; border-radius:14px;
            padding:16px 18px; box-shadow:0 1px 3px rgba(0,0,0,0.04); }
        div[data-testid="stMetricLabel"] p {
            color:#7A766E !important; font-size:.8rem; font-weight:600;
            text-transform:uppercase; letter-spacing:.4px; }
        [data-testid="stMetricValue"], [data-testid="stMetricValue"] * {
            color:#2F6F6A !important; font-weight:700; }
        .stButton>button, .stFormSubmitButton>button, .stDownloadButton>button,
        .stButton>button *, .stFormSubmitButton>button *, .stDownloadButton>button * {
            background:#2F6F6A; color:#FFFFFF !important; border:0;
            border-radius:10px; font-weight:600; }
        .stButton>button:hover, .stFormSubmitButton>button:hover,
        .stDownloadButton>button:hover { background:#255853 !important; }
        .marca-titulo { font-size:1.55rem; font-weight:700; color:#2A2A28 !important; margin-bottom:0; }
        .marca-sub { color:#7A766E !important; font-size:.9rem; margin-top:-4px; }
        .chip { display:inline-block; background:#EDE8DF; color:#5A5650 !important;
                border-radius:999px; padding:4px 12px; font-size:.78rem; font-weight:600; }
        #MainMenu {visibility:hidden;} footer {visibility:hidden;}
    </style>""", unsafe_allow_html=True)


# =============================================================================
# 3) BASE DE DATOS LOCAL (SQLite)
# =============================================================================
def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def ejecutar(sql: str, params: tuple = ()) -> None:
    conn = get_conn()
    try:
        conn.execute(sql, params)
        conn.commit()
    finally:
        conn.close()


def consultar(sql: str, params: tuple = ()) -> pd.DataFrame:
    conn = get_conn()
    try:
        return pd.read_sql_query(sql, conn, params=params)
    finally:
        conn.close()


def crear_tablas() -> None:
    conn = get_conn()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS clientes(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            empresa TEXT, contacto TEXT, telefono TEXT, correo TEXT, tipo TEXT, notas TEXT);
        CREATE TABLE IF NOT EXISTS obras(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT, cliente_id INTEGER, ubicacion TEXT, ingeniero TEXT,
            presupuesto REAL, fecha_inicio TEXT, fecha_fin TEXT, estatus TEXT);
        CREATE TABLE IF NOT EXISTS etapas(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, etapa TEXT, inicio TEXT, fin TEXT, estado TEXT, avance INTEGER);
        CREATE TABLE IF NOT EXISTS compras(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, fecha TEXT, categoria TEXT, descripcion TEXT, importe REAL,
            proveedor TEXT, comprador TEXT, comprobante TEXT, factura TEXT);
        CREATE TABLE IF NOT EXISTS requisiciones(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, folio TEXT, fecha TEXT, solicitante TEXT, material TEXT,
            cantidad REAL, unidad TEXT, proveedor TEXT, costo_estimado REAL, estatus TEXT);
        CREATE TABLE IF NOT EXISTS destajos(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, contratista TEXT, concepto TEXT,
            monto_contratado REAL, pagado REAL, avance INTEGER, estatus TEXT);
        CREATE TABLE IF NOT EXISTS bitacora(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, fecha TEXT, autor TEXT, nota TEXT);
        CREATE TABLE IF NOT EXISTS contratistas(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT, especialidad TEXT, telefono TEXT, correo TEXT);
        CREATE TABLE IF NOT EXISTS presupuesto(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, partida TEXT, concepto TEXT, monto REAL);
        CREATE TABLE IF NOT EXISTS proveedores(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT, agente TEXT, telefono TEXT, correo TEXT,
            cuenta TEXT, clabe TEXT, tarjeta TEXT);
        CREATE TABLE IF NOT EXISTS usuarios(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT, rol TEXT, telefono TEXT, correo TEXT,
            obra_id INTEGER, clave_hash TEXT, activo INTEGER DEFAULT 1);
        CREATE TABLE IF NOT EXISTS config(clave TEXT PRIMARY KEY, valor TEXT);
        CREATE TABLE IF NOT EXISTS contratos(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, nombre TEXT, contenido TEXT, fecha TEXT);
        CREATE TABLE IF NOT EXISTS contratos_contr(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contratista_id INTEGER, nombre TEXT, contenido TEXT, fecha TEXT);
        CREATE TABLE IF NOT EXISTS presupuesto_doc(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, nombre TEXT, contenido TEXT, fecha TEXT);
        CREATE TABLE IF NOT EXISTS abonos(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            obra_id INTEGER, fecha TEXT, concepto TEXT, monto REAL,
            metodo_pago TEXT, nota TEXT);
    """)
    # Migración: agrega la columna 'codigo' (clave de identificación) si falta
    for tabla in ["obras", "contratistas", "proveedores", "usuarios"]:
        try:
            conn.execute(f"ALTER TABLE {tabla} ADD COLUMN codigo TEXT")
        except Exception:
            pass
    try:
        conn.execute("ALTER TABLE obras ADD COLUMN contrato_link TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE obras ADD COLUMN presupuesto_link TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE contratistas ADD COLUMN monto_contratado REAL DEFAULT 0")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE contratistas ADD COLUMN contrato_link TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE compras ADD COLUMN metodo_pago TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE compras ADD COLUMN hora TEXT")
    except Exception:
        pass
    for col in ["banco", "beneficiario"]:
        try:
            conn.execute(f"ALTER TABLE proveedores ADD COLUMN {col} TEXT")
        except Exception:
            pass
    try:
        conn.execute("ALTER TABLE requisiciones ADD COLUMN prioridad TEXT DEFAULT 'Normal'")
    except Exception:
        pass
    for col in ["metodo_pago", "datos_bancarios", "banco_beneficiario"]:
        try:
            conn.execute(f"ALTER TABLE destajos ADD COLUMN {col} TEXT")
        except Exception:
            pass
    conn.commit()
    conn.close()


def inicializar_bd() -> None:
    """Crea las tablas. NO carga datos de ejemplo (app limpia)."""
    if "bd_lista" not in st.session_state:
        crear_tablas()
        os.makedirs(COMPROB_DIR, exist_ok=True)
        st.session_state.bd_lista = True


def borrar_todo() -> None:
    """Deja la app como recién adquirida: vacía todas las tablas y adjuntos."""
    for t in ["clientes", "obras", "etapas", "compras", "requisiciones",
              "destajos", "bitacora", "contratistas", "presupuesto", "proveedores", "usuarios"]:
        ejecutar(f"DELETE FROM {t}")
    if os.path.isdir(COMPROB_DIR):
        shutil.rmtree(COMPROB_DIR, ignore_errors=True)
    os.makedirs(COMPROB_DIR, exist_ok=True)


def obtener_obras() -> pd.DataFrame:
    return consultar("""SELECT o.*, c.empresa AS cliente
        FROM obras o LEFT JOIN clientes c ON o.cliente_id = c.id ORDER BY o.nombre""")


def obtener_clientes() -> pd.DataFrame:
    return consultar("SELECT * FROM clientes ORDER BY empresa")


def obtener_proveedores() -> pd.DataFrame:
    return consultar("SELECT * FROM proveedores ORDER BY nombre")


def _hash(clave: str) -> str:
    return hashlib.sha256((clave or "").encode("utf-8")).hexdigest()


def obtener_usuarios() -> pd.DataFrame:
    return consultar("""SELECT u.*, o.nombre AS obra
        FROM usuarios u LEFT JOIN obras o ON u.obra_id = o.id ORDER BY u.nombre""")


def opciones_clave(df, nombre_col="nombre"):
    """Construye opciones 'CLAVE · Nombre' y un mapa label->nombre para desplegables."""
    labels, mapa = [], {}
    for _, r in df.iterrows():
        cod = r["codigo"] if "codigo" in df.columns and pd.notna(r["codigo"]) and str(r["codigo"]).strip() else ""
        nom = str(r[nombre_col])
        label = f"{cod} · {nom}" if cod else nom
        labels.append(label)
        mapa[label] = nom
    return labels, mapa


# ---- Configuración (clave/valor) y Google Sheets ----
def cfg_get(clave, default=None):
    df = consultar("SELECT valor FROM config WHERE clave=?", (clave,))
    return df["valor"].iloc[0] if not df.empty else default


def cfg_set(clave, valor):
    ejecutar("INSERT INTO config(clave,valor) VALUES(?,?) "
             "ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor", (clave, valor))


def _secret(clave, default=None):
    try:
        return st.secrets[clave]
    except Exception:
        return default


def email_configurado() -> bool:
    cfg = _secret("email")
    return bool(cfg and cfg.get("smtp_host") and cfg.get("user") and cfg.get("password"))


def enviar_email(destinatario, asunto, cuerpo, pdf_bytes, nombre_archivo):
    """Envía un correo con el PDF adjunto vía SMTP (config en st.secrets['email'])."""
    import smtplib
    from email.message import EmailMessage
    cfg = _secret("email") or {}
    host = cfg.get("smtp_host")
    port = int(cfg.get("smtp_port", 587))
    user = cfg.get("user")
    pwd = cfg.get("password")
    remitente = cfg.get("from", user)
    msg = EmailMessage()
    msg["Subject"] = asunto
    msg["From"] = remitente
    msg["To"] = destinatario
    msg.set_content(cuerpo)
    if pdf_bytes:
        msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf",
                           filename=nombre_archivo)
    with smtplib.SMTP(host, port, timeout=30) as s:
        s.starttls()
        s.login(user, pwd)
        s.send_message(msg)


def bloque_enviar_reporte(pdf_bytes, titulo, nombre_archivo, resumen, clave):
    """Muestra opciones para enviar un reporte por WhatsApp o email (PDF)."""
    st.markdown("#### 📤 Enviar este reporte")
    cw, ce = st.columns(2)
    with cw:
        st.markdown("**Por WhatsApp**")
        num = st.text_input("WhatsApp (lada + número, solo dígitos)", value="52",
                            key=f"wa_{clave}")
        num_limpio = "".join(ch for ch in num if ch.isdigit())
        url = "https://wa.me/" + num_limpio + "?text=" + urllib.parse.quote(resumen)
        if len(num_limpio) >= 10:
            try:
                st.link_button("📲 Abrir WhatsApp con el mensaje", url)
            except Exception:
                st.markdown(f"[📲 Abrir WhatsApp con el mensaje]({url})")
        else:
            st.caption("Escribe lada + 10 dígitos para habilitar el envío.")
        st.caption("Adjunta el PDF descargado con un toque (WhatsApp no permite adjuntar "
                   "archivos automáticamente desde un enlace).")
    with ce:
        st.markdown("**Por correo electrónico**")
        correo = st.text_input("Correo del destinatario", key=f"mail_{clave}")
        if email_configurado():
            if st.button("📧 Enviar por email (con el PDF adjunto)", key=f"send_{clave}"):
                if not correo.strip():
                    st.warning("Escribe el correo del destinatario.")
                elif not pdf_bytes:
                    st.warning("No se pudo generar el PDF.")
                else:
                    try:
                        enviar_email(correo.strip(), titulo, resumen, pdf_bytes, nombre_archivo)
                        st.success(f"Reporte enviado a {correo.strip()}.")
                    except Exception as e:
                        st.error(f"No se pudo enviar el correo: {e}")
        else:
            asunto = urllib.parse.quote(titulo)
            cuerpo = urllib.parse.quote(resumen + "\n\n(Recuerda adjuntar el PDF descargado.)")
            try:
                st.link_button("📧 Abrir correo (adjunta el PDF)",
                               f"mailto:{correo}?subject={asunto}&body={cuerpo}")
            except Exception:
                st.markdown(f"[📧 Abrir correo](mailto:{correo}?subject={asunto}&body={cuerpo})")
            st.caption("Para que el PDF se envíe adjunto automáticamente, configura el correo "
                       "una sola vez (Ajustes → Secrets). Pídeme la plantilla.")


def _sheets_url():
    return _secret("gsheets_url") or cfg_get("gsheets_url")


def _credenciales():
    scopes = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    info = _secret("gcp_service_account")
    if info:
        return Credentials.from_service_account_info(dict(info), scopes=scopes)
    if os.path.exists(GCP_JSON):
        return Credentials.from_service_account_file(GCP_JSON, scopes=scopes)
    raise RuntimeError("No hay credenciales de Google configuradas.")


def _abrir_spreadsheet():
    gc = gspread.authorize(_credenciales())
    return gc.open_by_url(_sheets_url())


def sincronizar_a_sheets() -> int:
    """Sube (exporta) todas las tablas locales al Google Sheet."""
    sh = _abrir_spreadsheet()
    n = 0
    for t in TABLAS_SYNC:
        df = consultar(f"SELECT * FROM {t}")
        datos = [df.columns.tolist()] + df.astype(object).where(df.notna(), "").values.tolist()
        try:
            ws = sh.worksheet(t)
            ws.clear()
        except Exception:
            ws = sh.add_worksheet(title=t, rows=max(20, len(df) + 5),
                                  cols=max(5, len(df.columns) + 2))
        ws.update(datos)
        n += 1
    return n


def traer_de_sheets() -> int:
    """Trae (importa) los datos del Google Sheet y reemplaza las tablas locales."""
    sh = _abrir_spreadsheet()
    n = 0
    for t in TABLAS_SYNC:
        try:
            ws = sh.worksheet(t)
        except Exception:
            continue
        registros = ws.get_all_records()
        if not registros:
            continue
        ejecutar(f"DELETE FROM {t}")
        for r in registros:
            cols = list(r.keys())
            vals = [None if (v == "") else v for v in r.values()]
            ph = ",".join(["?"] * len(cols))
            try:
                ejecutar(f"INSERT INTO {t}({','.join(cols)}) VALUES({ph})", tuple(vals))
            except Exception:
                pass
        n += 1
    return n


def exportar_excel_bytes() -> bytes:
    """Genera un Excel con una hoja por cada tabla (base completa de la app)."""
    import io
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for t in TABLAS_SYNC:
            df = consultar(f"SELECT * FROM {t}")
            if df.empty:
                df = pd.DataFrame({"(sin registros)": []})
            df.to_excel(w, sheet_name=t[:31], index=False)
    return buf.getvalue()


def exportar_zip_csv_bytes() -> bytes:
    """Genera un .zip con un CSV por cada tabla (sin instalar nada extra)."""
    import io
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for t in TABLAS_SYNC:
            z.writestr(f"{t}.csv", consultar(f"SELECT * FROM {t}").to_csv(index=False))
    return buf.getvalue()


def obra_id_por_nombre(nombre) -> int:
    if not nombre:
        return -1
    df = consultar("SELECT id FROM obras WHERE nombre = ?", (nombre,))
    return int(df["id"].iloc[0]) if not df.empty else -1


# =============================================================================
# 4) ARCHIVOS ADJUNTOS (comprobantes y facturas)
# =============================================================================
def guardar_adjunto(archivo, prefijo: str):
    """Guarda un archivo subido en la carpeta de comprobantes y regresa su nombre."""
    if archivo is None:
        return None
    os.makedirs(COMPROB_DIR, exist_ok=True)
    ext = os.path.splitext(archivo.name)[1]
    nombre = f"{prefijo}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}{ext}"
    with open(os.path.join(COMPROB_DIR, nombre), "wb") as f:
        f.write(archivo.getbuffer())
    return nombre


def boton_descarga_adjunto(nombre_archivo: str, etiqueta: str, clave: str):
    """Muestra un botón para descargar/abrir un adjunto guardado."""
    if not nombre_archivo:
        return
    ruta = os.path.join(COMPROB_DIR, nombre_archivo)
    if os.path.exists(ruta):
        with open(ruta, "rb") as f:
            st.download_button(etiqueta, data=f.read(), file_name=nombre_archivo, key=clave)
    else:
        st.caption(f"({etiqueta}: archivo no encontrado)")


def guardar_contrato(obra_id, archivo):
    """Guarda el contrato (PDF) dentro de la base de datos, en base64."""
    import base64
    contenido = base64.b64encode(archivo.getbuffer()).decode("ascii")
    ejecutar("DELETE FROM contratos WHERE obra_id=?", (obra_id,))
    ejecutar("INSERT INTO contratos(obra_id,nombre,contenido,fecha) VALUES(?,?,?,?)",
             (obra_id, archivo.name, contenido, HOY.isoformat()))


def obtener_contrato(obra_id):
    df = consultar("SELECT nombre,contenido,fecha FROM contratos WHERE obra_id=?", (obra_id,))
    return df.iloc[0] if not df.empty else None


def _adjunto_existe(nombre_archivo) -> bool:
    """True si el adjunto realmente existe en disco (no solo un nombre guardado)."""
    if not isinstance(nombre_archivo, str) or not nombre_archivo.strip():
        return False
    return os.path.exists(os.path.join(COMPROB_DIR, nombre_archivo))


def guardar_contrato_contr(contratista_id, archivo):
    """Guarda el contrato (PDF) de un contratista dentro de la base de datos, en base64."""
    import base64
    contenido = base64.b64encode(archivo.getbuffer()).decode("ascii")
    ejecutar("DELETE FROM contratos_contr WHERE contratista_id=?", (contratista_id,))
    ejecutar("INSERT INTO contratos_contr(contratista_id,nombre,contenido,fecha) VALUES(?,?,?,?)",
             (contratista_id, archivo.name, contenido, HOY.isoformat()))


def obtener_contrato_contr(contratista_id):
    df = consultar("SELECT nombre,contenido,fecha FROM contratos_contr WHERE contratista_id=?",
                   (contratista_id,))
    return df.iloc[0] if not df.empty else None


def control_contratista(nombre):
    """Compara el monto contratado del contratista contra sus destajos."""
    cap_df = consultar("SELECT monto_contratado FROM contratistas WHERE nombre=?", (nombre,))
    cap = float(cap_df["monto_contratado"].iloc[0] or 0) if not cap_df.empty else 0.0
    d = consultar("SELECT COALESCE(SUM(monto_contratado),0) a, COALESCE(SUM(pagado),0) p "
                  "FROM destajos WHERE contratista=?", (nombre,))
    asignado = float(d["a"].iloc[0] or 0)
    pagado = float(d["p"].iloc[0] or 0)
    excedido = cap > 0 and (asignado > cap or pagado > cap)
    return {"cap": cap, "asignado": asignado, "pagado": pagado,
            "disponible": cap - asignado, "excedido": excedido}


def guardar_presupuesto_pdf(obra_id, archivo):
    """Guarda el presupuesto (PDF) dentro de la base de datos, en base64."""
    import base64
    contenido = base64.b64encode(archivo.getbuffer()).decode("ascii")
    ejecutar("DELETE FROM presupuesto_doc WHERE obra_id=?", (obra_id,))
    ejecutar("INSERT INTO presupuesto_doc(obra_id,nombre,contenido,fecha) VALUES(?,?,?,?)",
             (obra_id, archivo.name, contenido, HOY.isoformat()))


def obtener_presupuesto_pdf(obra_id):
    df = consultar("SELECT nombre,contenido,fecha FROM presupuesto_doc WHERE obra_id=?",
                   (obra_id,))
    return df.iloc[0] if not df.empty else None


def total_abonos(obra_id) -> float:
    df = consultar("SELECT COALESCE(SUM(monto),0) t FROM abonos WHERE obra_id=?", (obra_id,))
    return float(df["t"].iloc[0] or 0)


# =============================================================================
# 4B) GENERACIÓN DE PDF (un clic) con fpdf2
# =============================================================================
def _lat(s):
    """Texto seguro para PDF (latin-1)."""
    return str(s if s is not None else "").encode("latin-1", "replace").decode("latin-1")


# Paleta profesional basada en el logo (vino / negro)
PDF_PRIMARY = (123, 18, 18)     # vino del logo
PDF_DARK = (33, 33, 33)
PDF_GRAY = (120, 120, 120)
PDF_ROW = (247, 242, 242)
PDF_BORDER = (222, 210, 210)


class _ReportePDF(FPDF):
    def footer(self):
        self.set_y(-15)
        self.set_draw_color(*PDF_BORDER)
        self.set_line_width(0.2)
        self.line(10, self.get_y(), 200, self.get_y())
        self.set_y(-13)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(*PDF_GRAY)
        self.cell(95, 6, _lat(EMPRESA), align="L")
        self.cell(95, 6, _lat(f"Pagina {self.page_no()}"), align="R")


def _pdf_base(titulo: str, subtitulo: str = ""):
    pdf = _ReportePDF()
    pdf.set_auto_page_break(True, margin=18)
    pdf.add_page()
    # ---- Encabezado: logo + empresa + fecha ----
    if os.path.exists(LOGO_PATH):
        try:
            pdf.image(LOGO_PATH, x=10, y=10, w=26)
        except Exception:
            pass
    pdf.set_xy(40, 11)
    pdf.set_text_color(*PDF_PRIMARY)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(120, 6, _lat(EMPRESA))
    pdf.set_xy(40, 18)
    pdf.set_text_color(*PDF_GRAY)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(120, 5, _lat("Sistema de Control de Obras"))
    pdf.set_xy(120, 11)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*PDF_GRAY)
    pdf.cell(80, 5, _lat(f"Fecha: {HOY.isoformat()}"), align="R")
    # ---- Línea separadora ----
    pdf.set_y(31)
    pdf.set_draw_color(*PDF_PRIMARY)
    pdf.set_line_width(0.7)
    pdf.line(10, 31, 200, 31)
    # ---- Título del reporte ----
    pdf.set_y(37)
    pdf.set_x(10)
    pdf.set_text_color(*PDF_DARK)
    pdf.set_font("Helvetica", "B", 19)
    pdf.multi_cell(190, 10, _lat(titulo))
    if subtitulo:
        pdf.set_x(10)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*PDF_GRAY)
        pdf.multi_cell(190, 6, _lat(subtitulo))
    pdf.ln(2)
    pdf.set_text_color(*PDF_DARK)
    return pdf


def _pdf_titulo(pdf, t):
    pdf.ln(3)
    if pdf.get_y() > 265:
        pdf.add_page()
    y = pdf.get_y()
    pdf.set_fill_color(*PDF_PRIMARY)
    pdf.rect(10, y + 0.8, 2.5, 6, "F")
    pdf.set_xy(15, y)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*PDF_PRIMARY)
    pdf.multi_cell(185, 7, _lat(t))
    pdf.set_text_color(*PDF_DARK)
    pdf.set_x(10)


def _pdf_kv(pdf, pares):
    for k, v in pares:
        pdf.set_x(10)
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(*PDF_DARK)
        pdf.cell(55, 8, _lat(k), border=0)
        pdf.set_font("Helvetica", "", 11)
        pdf.multi_cell(135, 8, _lat(v))
    pdf.set_x(10)


def _pdf_kpis(pdf, items):
    """Dibuja tarjetas de indicadores (4 por fila)."""
    pdf.ln(2)
    x0, w, gap, h, perrow = 10, 45.25, 1.0, 19, 4
    base_y = pdf.get_y()
    for i, (label, value) in enumerate(items):
        fila, col = divmod(i, perrow)
        x = x0 + col * (w + gap)
        y = base_y + fila * (h + gap)
        pdf.set_fill_color(*PDF_ROW)
        pdf.set_draw_color(*PDF_BORDER)
        pdf.set_line_width(0.3)
        pdf.rect(x, y, w, h, "DF")
        pdf.set_xy(x + 3, y + 3)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*PDF_GRAY)
        pdf.cell(w - 6, 4, _lat(label))
        pdf.set_xy(x + 3, y + 9)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*PDF_PRIMARY)
        pdf.cell(w - 6, 7, _lat(value))
    filas_tot = (len(items) + perrow - 1) // perrow
    pdf.set_y(base_y + filas_tot * (h + gap) + 2)
    pdf.set_x(10)
    pdf.set_text_color(*PDF_DARK)


def _pdf_parrafo(pdf, texto, size=11, bold=False, color=None):
    pdf.set_x(10)
    pdf.set_font("Helvetica", "B" if bold else "", size)
    pdf.set_text_color(*(color if color else PDF_DARK))
    pdf.multi_cell(190, 8, _lat(texto))
    pdf.set_text_color(*PDF_DARK)
    pdf.set_x(10)


def _pdf_tabla(pdf, encabezados, filas, anchos):
    def _encabezado():
        pdf.set_x(10)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(*PDF_PRIMARY)
        pdf.set_text_color(255, 255, 255)
        for h, w in zip(encabezados, anchos):
            pdf.cell(w, 8, _lat(h), border=0, fill=True)
        pdf.ln(8)
    _encabezado()
    pdf.set_font("Helvetica", "", 9)
    pdf.set_draw_color(*PDF_BORDER)
    pdf.set_line_width(0.2)
    fill = False
    for fila in filas:
        if pdf.get_y() > 270:
            pdf.add_page()
            _encabezado()
            pdf.set_font("Helvetica", "", 9)
        pdf.set_x(10)
        pdf.set_text_color(*PDF_DARK)
        if fill:
            pdf.set_fill_color(*PDF_ROW)
        else:
            pdf.set_fill_color(255, 255, 255)
        for val, w in zip(fila, anchos):
            txt = _lat(str(val))
            maxc = max(4, int(w / 1.7))
            pdf.cell(w, 7, txt[:maxc], border="B", fill=True)
        pdf.ln(7)
        fill = not fill
    pdf.set_text_color(*PDF_DARK)


def pdf_compra(compra_id: int) -> bytes:
    c = consultar("SELECT * FROM compras WHERE id=?", (compra_id,)).iloc[0]
    obra = consultar("SELECT nombre,ubicacion FROM obras WHERE id=?", (int(c["obra_id"]),))
    on = obra["nombre"].iloc[0] if not obra.empty else ""
    ou = obra["ubicacion"].iloc[0] if not obra.empty else ""
    prov = consultar("SELECT * FROM proveedores WHERE nombre=?", (c["proveedor"],))
    pdf = _pdf_base("Comprobante de Compra")
    _pdf_kv(pdf, [("Obra", f"{on} - {ou}"), ("Fecha", c["fecha"]),
                  ("Categoria", c["categoria"]), ("Descripcion", c["descripcion"]),
                  ("Proveedor", c["proveedor"] or ""), ("Quien compra", c["comprador"] or ""),
                  ("Metodo de pago", c["metodo_pago"] if "metodo_pago" in c.index
                   and c["metodo_pago"] else "")])
    _pdf_parrafo(pdf, f"Importe: {pesos(c['importe'])} MXN", size=15, bold=True,
                 color=PDF_PRIMARY)
    if not prov.empty:
        p = prov.iloc[0]
        _pdf_titulo(pdf, "Datos del proveedor")
        _pdf_kv(pdf, [("Proveedor", p["nombre"]), ("Agente de ventas", p["agente"] or ""),
                      ("Contacto", f"{p['telefono'] or ''}  {p['correo'] or ''}"),
                      ("Cuenta", p["cuenta"] or ""), ("CLABE", p["clabe"] or ""),
                      ("Tarjeta", p["tarjeta"] or "")])
    pdf.ln(4)
    _pdf_parrafo(pdf, f"Generado por {EMPRESA} - {HOY.isoformat()}",
                 size=9, color=(122, 118, 110))
    return bytes(pdf.output())


def pdf_compras_lista(obra_id: int, ids: list, solicitante: str) -> bytes:
    """Documento imprimible de las compras a realizar (con fecha, hora y solicitante)."""
    o = consultar("SELECT nombre,ubicacion FROM obras WHERE id=?", (obra_id,))
    on = o["nombre"].iloc[0] if not o.empty else ""
    ahora = ahora_mx()
    pdf = _pdf_base("Compras a realizar", subtitulo=on)
    _pdf_kv(pdf, [("Fecha", ahora.strftime("%d/%m/%Y")),
                  ("Hora", ahora.strftime("%H:%M")),
                  ("Solicita la compra", solicitante or "")])
    if not ids:
        _pdf_parrafo(pdf, "No se seleccionaron compras.")
        return bytes(pdf.output())
    marcas = ",".join("?" * len(ids))
    comp = consultar(f"SELECT fecha,hora,categoria,descripcion,proveedor,metodo_pago,comprador,"
                     f"importe FROM compras WHERE id IN ({marcas}) ORDER BY fecha, id", ids)
    filas, total = [], 0.0
    for _, r in comp.iterrows():
        filas.append([r["fecha"] or "", r["hora"] or "", str(r["descripcion"] or ""),
                      str(r["proveedor"] or ""), str(r["metodo_pago"] or ""),
                      f"{pesos(r['importe'])}"])
        total += float(r["importe"] or 0)
    _pdf_titulo(pdf, "Detalle de compras")
    _pdf_tabla(pdf, ["Fecha", "Hora", "Concepto", "Proveedor", "Metodo", "Importe"],
               filas, [20, 14, 60, 44, 30, 22])
    _pdf_parrafo(pdf, f"TOTAL: {pesos(total)} MXN", size=14, bold=True, color=PDF_PRIMARY)
    pdf.ln(2)
    _pdf_parrafo(pdf, f"Generado por {EMPRESA} - {ahora.strftime('%d/%m/%Y %H:%M')}",
                 size=9, color=(122, 118, 110))
    return bytes(pdf.output())


def pdf_requisicion(req_id: int) -> bytes:
    r = consultar("SELECT * FROM requisiciones WHERE id=?", (req_id,)).iloc[0]
    obra = consultar("SELECT nombre,ubicacion FROM obras WHERE id=?", (int(r["obra_id"]),))
    on = obra["nombre"].iloc[0] if not obra.empty else ""
    ou = obra["ubicacion"].iloc[0] if not obra.empty else ""
    prioridad = r["prioridad"] if "prioridad" in r.index and r["prioridad"] else "Normal"
    pdf = _pdf_base(f"Requisicion de Material  -  {r['folio']}")
    _pdf_kv(pdf, [("Obra", f"{on} - {ou}"), ("Folio", r["folio"]), ("Fecha", r["fecha"]),
                  ("Solicitante", r["solicitante"] or ""), ("Material", r["material"]),
                  ("Cantidad", f"{r['cantidad']} {r['unidad']}"),
                  ("Proveedor", r["proveedor"] or ""),
                  ("Estatus", r["estatus"]), ("Prioridad", prioridad)])
    _pdf_parrafo(pdf, f"Costo estimado: {pesos(r['costo_estimado'])} MXN", size=15, bold=True,
                 color=PDF_PRIMARY)
    # Datos del proveedor para el pago
    prov = consultar("SELECT * FROM proveedores WHERE nombre=?", (r["proveedor"],))
    if not prov.empty:
        p = prov.iloc[0]

        def _g(col):
            return p[col] if col in prov.columns and p[col] else ""
        _pdf_titulo(pdf, "Datos del proveedor para el pago")
        _pdf_kv(pdf, [("Beneficiario de la cuenta", _g("beneficiario") or _g("nombre")),
                      ("Banco", _g("banco")),
                      ("No. de cuenta", _g("cuenta")),
                      ("CLABE interbancaria", _g("clabe")),
                      ("Tarjeta", _g("tarjeta")),
                      ("Correo (comprobante de pago)", _g("correo")),
                      ("Agente de ventas", _g("agente")),
                      ("Telefono", _g("telefono"))])
    pdf.ln(2)
    _pdf_parrafo(pdf, f"Generado por {EMPRESA} - {HOY.isoformat()}",
                 size=9, color=(122, 118, 110))
    return bytes(pdf.output())


def pdf_reporte(obra_id: int) -> bytes:
    k = calcular_kpis(obra_id)
    o = k["obra"]
    pdf = _pdf_base("Reporte de Obra", subtitulo=o["nombre"])
    _pdf_kv(pdf, [("Obra", o["nombre"]), ("Ubicacion", o["ubicacion"] or ""),
                  ("Responsable", o["ingeniero"] or "")])
    _pdf_kpis(pdf, [("% Avance general", f"{k['avance']}%"),
                    ("Presupuesto", f"{pesos(k['presupuesto'])}"),
                    ("Ejercido", f"{pesos(k['ejercido'])}"),
                    ("Dias restantes", str(k["dias"]))])
    et = consultar("SELECT etapa,estado,avance FROM etapas WHERE obra_id=?", (obra_id,))
    if not et.empty:
        _pdf_titulo(pdf, "Avance por etapa")
        _pdf_tabla(pdf, ["Etapa", "Estado", "%"], et.values.tolist(), [120, 50, 20])
    comp = consultar("SELECT categoria, SUM(importe) t FROM compras WHERE obra_id=? GROUP BY categoria",
                     (obra_id,))
    if not comp.empty:
        _pdf_titulo(pdf, "Compras por categoria")
        _pdf_tabla(pdf, ["Categoria", "Total"],
                   [[r[0], f"{pesos(r[1])}"] for r in comp.values.tolist()], [120, 70])
    de = consultar("SELECT contratista,monto_contratado,pagado FROM destajos WHERE obra_id=?",
                   (obra_id,))
    if not de.empty:
        _pdf_titulo(pdf, "Destajos")
        _pdf_tabla(pdf, ["Contratista", "Contratado", "Pagado"],
                   [[r[0], f"{pesos(r[1])}", f"{pesos(r[2])}"] for r in de.values.tolist()],
                   [100, 45, 45])
    return bytes(pdf.output())


# =============================================================================
# 5) ROLES Y PERMISOS
# =============================================================================
ROLES_ASIGNABLES = ["Departamento de Control", "Ingeniero de Obra", "Supervisor", "Cliente"]

PERMISOS = {
    "Administrador":           {"ver": True, "editar": True,  "crm": True,  "admin": True,  "usuarios": True,  "todas_obras": True},
    "Departamento de Control": {"ver": True, "editar": True,  "crm": True,  "admin": False, "usuarios": False, "todas_obras": True},
    "Ingeniero de Obra":       {"ver": True, "editar": True,  "crm": False, "admin": False, "usuarios": False, "todas_obras": False},
    "Supervisor":              {"ver": True, "editar": True,  "crm": False, "admin": False, "usuarios": False, "todas_obras": False},
    "Cliente":                 {"ver": True, "editar": False, "crm": False, "admin": False, "usuarios": False, "todas_obras": False},
}


def puede(rol: str, accion: str) -> bool:
    return PERMISOS.get(rol, {}).get(accion, False)


def verificar_clave_admin(clave_ingresada) -> bool:
    """True si la clave coincide con la del Administrador (o si no hay clave configurada)."""
    h = cfg_get("admin_pass_hash")
    if not h:
        return True
    return bool(clave_ingresada) and _hash(clave_ingresada) == h


def _candado_admin(clave_estado: str, etiqueta: str) -> bool:
    """Botón de lápiz que pide la clave del Administrador.
    Devuelve True solo cuando la edición está desbloqueada."""
    if st.session_state.get(f"unlock_{clave_estado}"):
        if st.button("🔒 Terminar edición", key=f"lock_{clave_estado}"):
            st.session_state[f"unlock_{clave_estado}"] = False
            st.rerun()
        return True
    if st.button(f"✏️ {etiqueta}", key=f"pencil_{clave_estado}"):
        st.session_state[f"ask_{clave_estado}"] = True
    if st.session_state.get(f"ask_{clave_estado}"):
        st.caption("Ingresa la clave del Administrador para ver y modificar los datos.")
        cl = st.text_input("Clave del Administrador", type="password",
                           key=f"clave_{clave_estado}")
        c1, c2 = st.columns(2)
        if c1.button("Desbloquear", key=f"unlock_btn_{clave_estado}"):
            if verificar_clave_admin(cl):
                st.session_state[f"unlock_{clave_estado}"] = True
                st.session_state[f"ask_{clave_estado}"] = False
                st.rerun()
            else:
                st.error("Clave del Administrador incorrecta.")
        if c2.button("Cancelar", key=f"cancel_{clave_estado}"):
            st.session_state[f"ask_{clave_estado}"] = False
            st.rerun()
    return False


def pantalla_login():
    """Pantalla de acceso. El Administrador entra libre; los demás con clave."""
    st.markdown("## 🔐 Acceso a Constructor PRO")
    st.caption("El Administrador entra libre. Los demás usuarios entran con la clave "
               "que les asignó el Administrador.")
    modo = st.radio("Tipo de acceso", ["Administrador (acceso libre)", "Usuario con clave"])
    if modo.startswith("Administrador"):
        admin_hash = cfg_get("admin_pass_hash")
        if admin_hash:
            pw = st.text_input("Clave del Administrador", type="password")
            if st.button("Entrar como Administrador"):
                if pw and _hash(pw) == admin_hash:
                    st.session_state.auth = {"nombre": "Administrador General",
                                             "rol": "Administrador", "obra_id": None}
                    st.rerun()
                else:
                    st.error("Clave de Administrador incorrecta.")
        else:
            if st.button("Entrar como Administrador"):
                st.session_state.auth = {"nombre": "Administrador General",
                                         "rol": "Administrador", "obra_id": None}
                st.rerun()
    else:
        usuarios = obtener_usuarios()
        if usuarios.empty:
            st.info("Aún no hay usuarios dados de alta. Entra como Administrador y créalos "
                    "en el apartado «Usuarios».")
        else:
            nombre = st.selectbox("Usuario", usuarios["nombre"].tolist())
            clave = st.text_input("Clave de acceso", type="password")
            if st.button("Entrar"):
                fila = usuarios[usuarios["nombre"] == nombre].iloc[0]
                if clave and _hash(clave) == (fila["clave_hash"] or ""):
                    st.session_state.auth = {
                        "nombre": nombre, "rol": fila["rol"],
                        "obra_id": int(fila["obra_id"]) if pd.notna(fila["obra_id"]) else None}
                    st.rerun()
                else:
                    st.error("Clave incorrecta.")


def obras_visibles(rol: str, obra_asignada) -> list:
    obras = obtener_obras()
    if obras.empty:
        return []
    if puede(rol, "todas_obras"):
        return obras["nombre"].tolist()
    if obra_asignada:
        return obras[obras["id"] == obra_asignada]["nombre"].tolist()
    return []


def selector_obra_trabajo(obra_id_actual, rol, etiqueta, key):
    """Para usuarios con acceso a todas las obras: muestra un menú desplegable de obra
    dentro de la sección y devuelve el obra_id elegido. Para los demás, deja la obra activa."""
    if not puede(rol, "todas_obras"):
        return obra_id_actual
    obras = obtener_obras()
    if obras.empty:
        return obra_id_actual
    labels, mapa = opciones_clave(obras)
    ids = [obra_id_por_nombre(mapa[l]) for l in labels]
    prev = st.session_state.get("obra_trabajo_id", obra_id_actual)
    if prev in ids:
        idx = ids.index(prev)
    elif obra_id_actual in ids:
        idx = ids.index(obra_id_actual)
    else:
        idx = 0
    sel = st.selectbox(etiqueta, labels, index=idx, key=key)
    elegido = obra_id_por_nombre(mapa[sel])
    st.session_state["obra_trabajo_id"] = elegido
    return elegido


def requiere_obra(obra_id: int) -> bool:
    """Muestra aviso si no hay obra activa. Devuelve True si SÍ hay obra."""
    if obra_id == -1:
        st.info("Todavía no hay obras. Da de alta una en el apartado «Obras» "
                "(y un cliente en «CRM»). Luego selecciónala arriba en «Obra activa».")
        return False
    return True


# =============================================================================
# 6) KPIs
# =============================================================================
def calcular_kpis(obra_id: int) -> dict:
    obra = consultar("SELECT * FROM obras WHERE id = ?", (obra_id,)).iloc[0]
    etapas = consultar("SELECT avance FROM etapas WHERE obra_id = ?", (obra_id,))
    avance = round(etapas["avance"].mean(), 1) if not etapas.empty else 0.0
    ejercido = consultar("SELECT COALESCE(SUM(importe),0) AS s FROM compras WHERE obra_id=?",
                         (obra_id,))["s"].iloc[0]
    presupuesto = float(obra["presupuesto"]) if obra["presupuesto"] else 0.0
    pct = round(ejercido / presupuesto * 100, 1) if presupuesto else 0
    fin = datetime.strptime(obra["fecha_fin"], "%Y-%m-%d").date() if obra["fecha_fin"] else HOY
    dias = (fin - HOY).days
    return {"avance": avance, "presupuesto": presupuesto, "ejercido": float(ejercido),
            "pct": pct, "dias": dias, "obra": obra}


# =============================================================================
# 7) GRÁFICAS
# =============================================================================
def _mate(fig, alto=380):
    fig.update_layout(height=alto, margin=dict(l=10, r=10, t=40, b=10),
                      paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="#FFFFFF",
                      font=dict(color="#3A3A37", size=13),
                      legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
    fig.update_xaxes(gridcolor="#ECE8E0", zeroline=False)
    fig.update_yaxes(gridcolor="#ECE8E0", zeroline=False)
    return fig


def grafica_gauge(avance):
    fig = go.Figure(go.Indicator(
        mode="gauge+number", value=avance, number={"suffix": "%"},
        gauge={"axis": {"range": [0, 100]}, "bar": {"color": COLOR_PRIMARIO},
               "steps": [{"range": [0, 50], "color": "#F0EBE2"},
                         {"range": [50, 80], "color": "#E3DDD0"},
                         {"range": [80, 100], "color": "#D7E3D8"}]},
        title={"text": "Avance general"}))
    return _mate(fig, 300)


def grafica_gantt(df):
    if df.empty:
        return _mate(go.Figure().add_annotation(text="Sin etapas", showarrow=False), 300)
    df = df.copy()
    df["inicio"] = pd.to_datetime(df["inicio"]); df["fin"] = pd.to_datetime(df["fin"])
    mapa = {"Completada": COLOR_OK, "En proceso": COLOR_PRIMARIO, "Por iniciar": COLOR_ACENTO}
    fig = px.timeline(df, x_start="inicio", x_end="fin", y="etapa", color="estado",
                      color_discrete_map=mapa, hover_data={"avance": True},
                      title="Cronograma de obra (Gantt)")
    fig.update_yaxes(autorange="reversed")
    fig.add_vline(x=pd.Timestamp(HOY), line_width=2, line_dash="dash", line_color=COLOR_ALERTA)
    return _mate(fig, 400)


def grafica_avance_etapas(df):
    if df.empty:
        return _mate(go.Figure().add_annotation(text="Sin etapas", showarrow=False), 300)
    fig = px.bar(df, x="avance", y="etapa", orientation="h", title="% de avance por etapa",
                 text="avance", color_discrete_sequence=[COLOR_PRIMARIO])
    fig.update_traces(texttemplate="%{text}%", textposition="outside")
    fig.update_layout(xaxis_title="% avance", yaxis_title="", xaxis_range=[0, 110])
    fig.update_yaxes(autorange="reversed")
    return _mate(fig, 360)


def grafica_destajos(df):
    if df.empty:
        return _mate(go.Figure().add_annotation(text="Sin destajos", showarrow=False), 300)
    fig = go.Figure()
    fig.add_trace(go.Bar(x=df["contratista"], y=df["monto_contratado"], name="Contratado",
                         marker_color=COLOR_ACENTO))
    fig.add_trace(go.Bar(x=df["contratista"], y=df["pagado"], name="Pagado",
                         marker_color=COLOR_PRIMARIO))
    fig.update_layout(barmode="group", title="Destajos: contratado vs pagado", yaxis_title="MXN")
    return _mate(fig, 360)


def grafica_compras_categoria(df):
    if df.empty:
        return _mate(go.Figure().add_annotation(text="Sin compras registradas", showarrow=False), 300)
    g = df.groupby("categoria", as_index=False)["importe"].sum().sort_values("importe")
    g["_etq"] = g["importe"].map(pesos)
    fig = px.bar(g, x="importe", y="categoria", orientation="h",
                 title="Compras por categoría", text="_etq",
                 color_discrete_sequence=[COLOR_PRIMARIO])
    fig.update_traces(textposition="outside")
    fig.update_layout(xaxis_title="MXN", yaxis_title="")
    return _mate(fig, 360)


def grafica_presupuesto_partidas(df):
    if df.empty:
        return _mate(go.Figure().add_annotation(text="Sin presupuesto cargado", showarrow=False), 300)
    g = df.groupby("partida", as_index=False)["monto"].sum().sort_values("monto")
    g["_etq"] = g["monto"].map(pesos)
    fig = px.bar(g, x="monto", y="partida", orientation="h",
                 title="Presupuesto por partida", text="_etq",
                 color_discrete_sequence=[COLOR_PRIMARIO])
    fig.update_traces(textposition="outside")
    fig.update_layout(xaxis_title="MXN", yaxis_title="")
    return _mate(fig, 360)


# =============================================================================
# 8) VISTAS
# =============================================================================
def pdf_pagos_semana(obra_id: int) -> bytes:
    """Reporte semanal: suma de todos los destajos por pagar de la obra."""
    o = consultar("SELECT nombre,codigo FROM obras WHERE id=?", (obra_id,))
    on = o["nombre"].iloc[0] if not o.empty else ""
    cod = ""
    if not o.empty and "codigo" in o.columns and pd.notna(o["codigo"].iloc[0]):
        cod = str(o["codigo"].iloc[0])
    dest = consultar("SELECT contratista,concepto,monto_contratado,pagado,metodo_pago,"
                     "datos_bancarios,banco_beneficiario FROM destajos WHERE obra_id=?", (obra_id,))
    lunes = HOY - timedelta(days=HOY.weekday())
    domingo = lunes + timedelta(days=6)
    pdf = _pdf_base("Reporte semanal de pagos a destajos")
    _pdf_kv(pdf, [("Obra", f"{cod + ' - ' if cod else ''}{on}"),
                  ("Semana", f"{lunes.isoformat()} al {domingo.isoformat()}")])

    def _t(v):
        return str(v) if (pd.notna(v) and str(v).strip()) else "-"

    _pdf_titulo(pdf, "Destajos a pagar esta semana")
    hay, total = False, 0.0
    for _, r in dest.iterrows():
        saldo = float(r["monto_contratado"]) - float(r["pagado"])
        if saldo <= 0.009:
            continue
        hay = True
        total += saldo
        _pdf_parrafo(pdf, f"{r['contratista']}  -  {r['concepto']}", size=11, bold=True)
        _pdf_kv(pdf, [("Metodo de pago", _t(r["metodo_pago"])),
                      ("No. tarjeta / CLABE / cuenta", _t(r["datos_bancarios"])),
                      ("Banco y beneficiario", _t(r["banco_beneficiario"])),
                      ("A pagar", f"{pesos(saldo)} MXN")])
        pdf.ln(1)
    if not hay:
        _pdf_parrafo(pdf, "No hay saldos pendientes de pago.")
    _pdf_parrafo(pdf, f"TOTAL A PAGAR: {pesos(total)} MXN", size=15, bold=True,
                 color=PDF_PRIMARY)
    pdf.ln(2)
    _pdf_parrafo(pdf, f"Generado por {EMPRESA} - {HOY.isoformat()}", size=9,
                 color=(122, 118, 110))
    return bytes(pdf.output())


def vista_dashboard(obra_id: int):
    if not requiere_obra(obra_id):
        return
    k = calcular_kpis(obra_id); o = k["obra"]
    cliente = obtener_obras().set_index("id").loc[obra_id, "cliente"]
    st.subheader("📊 Panel general")
    st.caption(f"Cliente: {cliente}  ·  Ubicación: {o['ubicacion']}  ·  Responsable: {o['ingeniero']}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("% Avance general", f"{k['avance']}%")
    c2.metric("Presupuesto", f"{pesos(k['presupuesto'])}")
    c3.metric("Ejercido (compras)", f"{pesos(k['ejercido'])}", f"{k['pct']}% del total")
    c4.metric("Días restantes", f"{k['dias']} días", "Vencida" if k["dias"] < 0 else "En tiempo")

    # Resumen del Control Financiero
    recibido = total_abonos(obra_id)
    pagado_dest = consultar("SELECT COALESCE(SUM(pagado),0) p FROM destajos WHERE obra_id=?",
                            (obra_id,))["p"].iloc[0]
    egresos = k["ejercido"] + float(pagado_dest or 0)
    balance = recibido - egresos
    st.markdown("##### 💰 Control financiero")
    f1, f2, f3 = st.columns(3)
    f1.metric("Recibido (abonos)", f"{pesos(recibido)}")
    f2.metric("Egresos (compras + destajos)", f"{pesos(egresos)}")
    f3.metric("Balance", f"{pesos(balance)}", "A favor" if balance >= 0 else "En contra")
    st.markdown("---")
    etapas = consultar("SELECT * FROM etapas WHERE obra_id=?", (obra_id,))
    destajos = consultar("SELECT * FROM destajos WHERE obra_id=?", (obra_id,))
    compras = consultar("SELECT * FROM compras WHERE obra_id=?", (obra_id,))
    g1, g2 = st.columns([1, 2])
    with g1: st.plotly_chart(grafica_gauge(k["avance"]), width="stretch", key="plt_1")
    with g2: st.plotly_chart(grafica_avance_etapas(etapas), width="stretch", key="plt_2")
    st.plotly_chart(grafica_gantt(etapas), width="stretch", key="plt_3")
    g3, g4 = st.columns(2)
    with g3: st.plotly_chart(grafica_destajos(destajos), width="stretch", key="plt_4")
    with g4: st.plotly_chart(grafica_compras_categoria(compras), width="stretch", key="plt_5")


def comprobante_compra_html(compra_id: int) -> str:
    """Genera un comprobante imprimible (PDF vía navegador) de una compra."""
    c = consultar("SELECT * FROM compras WHERE id=?", (compra_id,)).iloc[0]
    obra = consultar("SELECT nombre,ubicacion FROM obras WHERE id=?", (int(c["obra_id"]),))
    obra_nombre = obra["nombre"].iloc[0] if not obra.empty else ""
    obra_ubic = obra["ubicacion"].iloc[0] if not obra.empty else ""
    prov = consultar("SELECT * FROM proveedores WHERE nombre=?", (c["proveedor"],))
    datos_banco = ""
    if not prov.empty:
        p = prov.iloc[0]
        datos_banco = f"""
        <h3>Datos del proveedor</h3>
        <table>
          <tr><th>Proveedor</th><td>{p['nombre']}</td></tr>
          <tr><th>Agente de ventas</th><td>{p['agente'] or ''}</td></tr>
          <tr><th>Contacto</th><td>{p['telefono'] or ''} &nbsp; {p['correo'] or ''}</td></tr>
          <tr><th>Cuenta</th><td>{p['cuenta'] or ''}</td></tr>
          <tr><th>CLABE</th><td>{p['clabe'] or ''}</td></tr>
          <tr><th>Tarjeta</th><td>{p['tarjeta'] or ''}</td></tr>
        </table>"""
    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">
    <title>Comprobante de compra</title><style>
      body{{font-family:'Segoe UI',Arial,sans-serif;color:#2A2A28;margin:40px;max-width:720px;}}
      h1{{color:#2F6F6A;border-bottom:3px solid #C9842B;padding-bottom:8px;}}
      h3{{color:#2F6F6A;margin-top:22px;}}
      table{{border-collapse:collapse;width:100%;margin-top:8px;font-size:14px;}}
      th{{background:#F0EBE2;text-align:left;padding:9px;width:38%;}}
      td{{border-bottom:1px solid #E6E2DA;padding:9px;}}
      .total{{font-size:1.6rem;color:#2F6F6A;font-weight:700;margin-top:18px;}}
      .pie{{margin-top:30px;color:#7A766E;font-size:12px;}}
      @media print {{ button{{display:none;}} }}
    </style></head><body>
    <button onclick="window.print()" style="background:#2F6F6A;color:#fff;border:0;
      padding:10px 20px;border-radius:8px;cursor:pointer;">🖨️ Imprimir / Guardar PDF</button>
    <h1>Comprobante de Compra</h1>
    <h3>Datos de la compra</h3>
    <table>
      <tr><th>Obra</th><td>{obra_nombre} — {obra_ubic}</td></tr>
      <tr><th>Fecha</th><td>{c['fecha']}</td></tr>
      <tr><th>Categoría</th><td>{c['categoria']}</td></tr>
      <tr><th>Descripción</th><td>{c['descripcion']}</td></tr>
      <tr><th>Proveedor</th><td>{c['proveedor'] or ''}</td></tr>
      <tr><th>Quién realiza la compra</th><td>{c['comprador'] or ''}</td></tr>
    </table>
    <p class="total">Importe: {pesos(c['importe'])} MXN</p>
    {datos_banco}
    <p class="pie">Generado por SONA CONSTRUCTORES DEL MAYAB · {HOY.isoformat()}</p>
    </body></html>"""


def vista_compras(obra_id: int, rol: str, usuario: str):
    st.subheader("🛒 Compras y gastos diarios")
    if not requiere_obra(obra_id):
        return
    if puede(rol, "todas_obras"):
        obra_id = selector_obra_trabajo(obra_id, rol,
                                        "🏢 Obra a la que se cargará la compra", "compras_obra_sel")
    obra_nombre = consultar("SELECT nombre FROM obras WHERE id=?", (obra_id,))["nombre"].iloc[0]
    st.caption(f"Las compras se cargarán a la obra: **{obra_nombre}**")

    proveedores = obtener_proveedores()
    prov_labels, prov_map = opciones_clave(proveedores) if not proveedores.empty else ([], {})
    usuarios_df = obtener_usuarios()
    if not usuarios_df.empty:
        asign = usuarios_df[usuarios_df["obra_id"] == obra_id]
        compradores = asign["nombre"].tolist()
    else:
        compradores = []
    compras = consultar("SELECT * FROM compras WHERE obra_id=? ORDER BY fecha DESC, id DESC",
                        (obra_id,))
    if not compras.empty:
        c1, c2, c3 = st.columns(3)
        c1.metric("Compras registradas", len(compras))
        c2.metric("Total comprado", f"{pesos(compras['importe'].sum())}")
        hoy_total = compras[compras["fecha"] == HOY.isoformat()]["importe"].sum()
        c3.metric("Comprado hoy", f"{pesos(hoy_total)}")
        st.plotly_chart(grafica_compras_categoria(compras), width="stretch", key="plt_6")

    if puede(rol, "editar"):
        # Alta rápida de proveedor (fuera del formulario para refrescar el listado)
        with st.expander("➕ Registrar un proveedor nuevo"):
            with st.form("form_prov_rapido", clear_on_submit=True):
                col1, col2 = st.columns(2)
                with col1:
                    p_codigo = st.text_input("Clave del proveedor")
                    p_nombre = st.text_input("Proveedor (empresa)")
                    p_agente = st.text_input("Agente de ventas")
                    p_tel = st.text_input("Teléfono de contacto")
                    p_correo = st.text_input("Correo")
                with col2:
                    p_cuenta = st.text_input("Número de cuenta")
                    p_clabe = st.text_input("CLABE interbancaria")
                    p_tarjeta = st.text_input("Tarjeta")
                if st.form_submit_button("💾 Guardar proveedor") and p_nombre.strip():
                    ejecutar("INSERT INTO proveedores(codigo,nombre,agente,telefono,correo,cuenta,"
                             "clabe,tarjeta) VALUES(?,?,?,?,?,?,?,?)",
                             (p_codigo.strip(), p_nombre.strip(), p_agente, p_tel, p_correo,
                              p_cuenta, p_clabe, p_tarjeta))
                    st.success(f"Proveedor «{p_nombre}» guardado."); st.rerun()

        partidas_pres = consultar("SELECT partida, concepto FROM presupuesto WHERE obra_id=? "
                                  "ORDER BY partida", (obra_id,))
        opciones_desc = []
        for _, rp in partidas_pres.iterrows():
            etq = str(rp["partida"] or "").strip()
            if rp["concepto"]:
                etq = f"{etq} · {rp['concepto']}" if etq else str(rp["concepto"])
            if etq:
                opciones_desc.append(etq)
        st.markdown("#### ➕ Registrar una compra")
        with st.form("form_compra", clear_on_submit=True):
            col1, col2, col3 = st.columns(3)
            with col1:
                fecha = st.date_input("Fecha de la compra", HOY)
                categoria = st.selectbox("Categoría del gasto", CATEGORIAS_COMPRA)
            with col2:
                if opciones_desc:
                    desc_part = st.selectbox("Descripción / concepto (partida del presupuesto)",
                                             ["(elegir partida)"] + opciones_desc)
                    desc_otro = st.text_input("...o escribe otra descripción")
                else:
                    desc_part = "(elegir partida)"
                    desc_otro = st.text_input("Descripción / concepto")
                if prov_labels:
                    prov_label = st.selectbox("Proveedor (del catálogo)", prov_labels)
                    proveedor = prov_map[prov_label]
                else:
                    proveedor = st.text_input("Proveedor (regístralo arriba o en «Proveedores»)")
            with col3:
                importe = st.number_input("Importe ($ MXN)", min_value=0.0, step=10.0, format="%.2f")
                if compradores:
                    comprador = st.selectbox("Quién realiza la compra", compradores)
                else:
                    comprador = st.text_input("Quién realiza la compra", value=usuario)
                metodo_pago = st.selectbox("Método de pago", METODOS_PAGO)
            colA, colB = st.columns(2)
            with colA:
                comprobante = st.file_uploader("Comprobante (foto o PDF)",
                                               type=["png", "jpg", "jpeg", "pdf"])
            with colB:
                factura = st.file_uploader("Factura (PDF o XML)",
                                           type=["pdf", "xml", "png", "jpg", "jpeg"])
            enviar = st.form_submit_button("💾 Guardar compra")
        descripcion = desc_otro.strip() if desc_otro.strip() else (
            desc_part if desc_part != "(elegir partida)" else "")
        if enviar and descripcion.strip():
            nom_comp = guardar_adjunto(comprobante, "comp")
            nom_fact = guardar_adjunto(factura, "fact")
            ejecutar("INSERT INTO compras(obra_id,fecha,categoria,descripcion,importe,"
                     "proveedor,comprador,comprobante,factura,metodo_pago,hora) "
                     "VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                     (obra_id, fecha.isoformat(), categoria, descripcion.strip(), importe,
                      proveedor, comprador.strip(), nom_comp, nom_fact, metodo_pago,
                      ahora_mx().strftime("%H:%M")))
            st.success("Compra registrada y cargada a la obra.")
            st.rerun()
        elif enviar:
            st.warning("Elige una partida del presupuesto o escribe la descripción de la compra.")
    else:
        st.info("Tu rol (Cliente) es de solo lectura.")

    st.markdown("#### Historial de compras")
    if compras.empty:
        st.caption("Aún no hay compras registradas para esta obra.")
        return
    cols_hist = ["fecha", "categoria", "descripcion", "proveedor", "comprador", "metodo_pago",
                 "importe"]
    cols_hist = [c for c in cols_hist if c in compras.columns]
    tabla = compras[cols_hist].copy()
    tabla["importe"] = tabla["importe"].map(lambda x: f"{pesos(x)}")
    st.dataframe(tabla.rename(columns={"fecha": "Fecha", "categoria": "Categoría",
                 "descripcion": "Descripción", "proveedor": "Proveedor",
                 "comprador": "Comprador", "metodo_pago": "Método de pago",
                 "importe": "Importe"}),
                 width="stretch", hide_index=True)

    # ---- Imprimir la o las compras a realizar ----
    st.markdown("#### 🖨️ Imprimir compras a realizar")
    st.caption("Selecciona una o varias compras; el documento incluye fecha, hora y quién solicita.")
    opciones_imp = {
        f"{r['fecha']} · {r['descripcion']} · {pesos(r['importe'])}": int(r["id"])
        for _, r in compras.iterrows()}
    sel_imp = st.multiselect("Compras a imprimir", list(opciones_imp.keys()),
                             default=list(opciones_imp.keys()), key="imp_compras_sel")
    ids_imp = [opciones_imp[s] for s in sel_imp]
    if FPDF_OK:
        if ids_imp:
            st.download_button("📄 Descargar documento de compras (PDF)",
                               data=pdf_compras_lista(obra_id, ids_imp, usuario),
                               file_name=f"Compras_a_realizar_{HOY.isoformat()}.pdf",
                               mime="application/pdf", key="dl_compras_lista")
        else:
            st.caption("Selecciona al menos una compra para generar el documento.")
    else:
        st.caption("Para el PDF instala una vez: python -m pip install fpdf2")

    st.markdown("#### Comprobante PDF, adjuntos y factura por compra")
    con_adjuntos = compras[compras.apply(
        lambda c: _adjunto_existe(c["comprobante"]) or _adjunto_existe(c["factura"]), axis=1)]
    if con_adjuntos.empty:
        st.caption("Aún no hay compras con comprobante o factura adjunta.")
        return
    st.caption("Solo se muestran las compras que tienen comprobante o factura subidos:")
    for _, c in con_adjuntos.iterrows():
        st.write(f"**{c['fecha']}** · {c['categoria']} · {c['descripcion']} ({pesos(c['importe'])})")
        tiene_comp = _adjunto_existe(c["comprobante"])
        tiene_fact = _adjunto_existe(c["factura"])
        cols = st.columns(1 + int(tiene_comp) + int(tiene_fact))
        i = 0
        with cols[i]:
            if FPDF_OK:
                st.download_button("⬇️ Comprobante PDF", data=pdf_compra(int(c["id"])),
                                   file_name=f"Compra_{int(c['id'])}.pdf",
                                   mime="application/pdf", key=f"pdf_{c['id']}")
            else:
                st.download_button("🖨️ Comprobante (HTML)",
                                   data=comprobante_compra_html(int(c["id"])),
                                   file_name=f"Compra_{int(c['id'])}.html",
                                   mime="text/html", key=f"pdf_{c['id']}")
        if tiene_comp:
            i += 1
            with cols[i]:
                boton_descarga_adjunto(c["comprobante"], "📎 Adjunto", f"comp_{c['id']}")
        if tiene_fact:
            i += 1
            with cols[i]:
                boton_descarga_adjunto(c["factura"], "🧾 Factura", f"fact_{c['id']}")
        st.markdown("<hr style='margin:4px 0;border:none;border-top:1px solid #E6E2DA;'>",
                    unsafe_allow_html=True)


def vista_crm(rol: str):
    st.subheader("🤝 CRM · Clientes y prospectos")
    if not puede(rol, "crm"):
        st.warning("El módulo CRM es exclusivo del Administrador."); return
    clientes = obtener_clientes(); obras = obtener_obras()
    c1, c2, c3 = st.columns(3)
    c1.metric("Clientes totales", len(clientes))
    c2.metric("Prospectos", int((clientes["tipo"] == "Prospecto").sum()) if not clientes.empty else 0)
    c3.metric("Cartera de obras", f"{pesos(obras['presupuesto'].sum() if not obras.empty else 0)}")
    if clientes.empty:
        st.caption("Aún no hay clientes. Registra el primero abajo.")
    else:
        st.markdown("#### Cartera por cliente")
        for _, cli in clientes.iterrows():
            obras_cli = obras[obras["cliente"] == cli["empresa"]] if not obras.empty else pd.DataFrame()
            tot = obras_cli["presupuesto"].sum() if not obras_cli.empty else 0
            with st.expander(f"🏢 {cli['empresa']}  ·  {cli['tipo']}  ·  {pesos(tot)}"):
                st.write(f"Contacto: {cli['contacto']}  ·  📞 {cli['telefono']}  ·  ✉️ {cli['correo']}")
                if cli["notas"]:
                    st.write(f"*{cli['notas']}*")
                if not obras_cli.empty:
                    st.dataframe(obras_cli[["nombre", "ubicacion", "presupuesto", "estatus"]],
                                 width="stretch", hide_index=True)
    st.markdown("---")
    st.markdown("#### Registrar nuevo cliente / prospecto")
    with st.form("form_cliente", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            empresa = st.text_input("Empresa"); contacto = st.text_input("Contacto")
            telefono = st.text_input("Teléfono")
        with col2:
            correo = st.text_input("Correo"); tipo = st.selectbox("Tipo", TIPO_CLIENTE)
            notas = st.text_input("Notas")
        if st.form_submit_button("➕ Guardar cliente") and empresa.strip():
            ejecutar("INSERT INTO clientes(empresa,contacto,telefono,correo,tipo,notas) "
                     "VALUES(?,?,?,?,?,?)", (empresa.strip(), contacto, telefono, correo, tipo, notas))
            st.success(f"Cliente «{empresa}» guardado."); st.rerun()

    if not clientes.empty:
        st.markdown("---")
        st.markdown("#### ✏️ Modificar un cliente")
        st.caption("Si registraste un cliente con datos incompletos, aquí puedes completarlos "
                   "o corregirlos después.")
        emp_sel = st.selectbox("Selecciona el cliente a modificar",
                               clientes["empresa"].tolist(), key="crm_edit_sel")
        c = clientes[clientes["empresa"] == emp_sel].iloc[0]
        with st.form("form_cliente_editar"):
            col1, col2 = st.columns(2)
            with col1:
                e_empresa = st.text_input("Empresa", value=c["empresa"] or "")
                e_contacto = st.text_input("Contacto", value=c["contacto"] or "")
                e_telefono = st.text_input("Teléfono", value=c["telefono"] or "")
            with col2:
                e_correo = st.text_input("Correo", value=c["correo"] or "")
                tipo_idx = TIPO_CLIENTE.index(c["tipo"]) if c["tipo"] in TIPO_CLIENTE else 0
                e_tipo = st.selectbox("Tipo", TIPO_CLIENTE, index=tipo_idx)
                e_notas = st.text_input("Notas", value=c["notas"] or "")
            if st.form_submit_button("💾 Guardar cambios") and e_empresa.strip():
                ejecutar("UPDATE clientes SET empresa=?, contacto=?, telefono=?, correo=?, "
                         "tipo=?, notas=? WHERE id=?",
                         (e_empresa.strip(), e_contacto, e_telefono, e_correo, e_tipo,
                          e_notas, int(c["id"])))
                st.success(f"Datos de «{e_empresa}» actualizados."); st.rerun()


def vista_respaldo(rol: str):
    st.subheader("💾 Respaldo / Exportar a Excel")
    if not puede(rol, "admin"):
        st.warning("Solo el Administrador puede exportar la base completa."); return
    st.write("Descarga una copia de **toda la información cargada en la app** "
             "(obras, compras, proveedores, contratistas, destajos, usuarios, etc.). "
             "En el Excel, cada tabla queda en su propia hoja.")
    resumen = pd.DataFrame([(t, len(consultar(f"SELECT * FROM {t}"))) for t in TABLAS_SYNC],
                           columns=["Tabla", "Registros"])
    st.dataframe(resumen, width="stretch", hide_index=True)

    col1, col2 = st.columns(2)
    with col1:
        if OPENPYXL_OK:
            st.download_button("⬇️ Descargar TODO en Excel (.xlsx)",
                               data=exportar_excel_bytes(),
                               file_name=f"Respaldo_SONA_{HOY.isoformat()}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument."
                                    "spreadsheetml.sheet")
        else:
            st.info("Para el Excel instala una vez:  python -m pip install openpyxl")
    with col2:
        st.download_button("⬇️ Descargar TODO en CSV (.zip)",
                           data=exportar_zip_csv_bytes(),
                           file_name=f"Respaldo_SONA_{HOY.isoformat()}.zip",
                           mime="application/zip")
    st.caption("Sugerencia: guarda un respaldo cada semana. El archivo «cob_data.db» "
               "que está junto a la app también es tu base de datos completa.")


def vista_sheets(rol: str):
    st.subheader("☁️ Vincular con Google Sheets")
    if not puede(rol, "admin"):
        st.warning("Solo el Administrador puede configurar Google Sheets."); return

    if not GSHEETS_OK:
        st.error("Faltan librerías. Instala una sola vez en la terminal:  "
                 "python -m pip install gspread google-auth")
        return

    creds_ok = os.path.exists(GCP_JSON) or bool(_secret("gcp_service_account"))
    url_ok = bool(_sheets_url())
    c1, c2 = st.columns(2)
    c1.metric("Credenciales", "Cargadas ✅" if creds_ok else "Faltan ❌")
    c2.metric("Hoja vinculada", "Configurada ✅" if url_ok else "Falta ❌")

    with st.expander("📋 Cómo configurarlo (una sola vez)", expanded=not (creds_ok and url_ok)):
        st.markdown(
            "1. Entra a **https://console.cloud.google.com** y crea un proyecto.\n"
            "2. Activa **Google Sheets API** y **Google Drive API**.\n"
            "3. Crea una **Cuenta de servicio** y genera una **llave JSON** (la descarga).\n"
            "4. Crea un Google Sheet en blanco y **compártelo** (botón Compartir) con el "
            "correo de la cuenta de servicio (termina en `gserviceaccount.com`) como **Editor**.\n"
            "5. Sube aquí abajo el archivo JSON y pega la **URL** del Google Sheet.")

    st.markdown("#### 1) Cargar credenciales (archivo JSON de la cuenta de servicio)")
    jsonf = st.file_uploader("Archivo .json de la cuenta de servicio", type=["json"])
    if jsonf is not None and st.button("💾 Guardar credenciales"):
        with open(GCP_JSON, "wb") as f:
            f.write(jsonf.getbuffer())
        st.success("Credenciales guardadas."); st.rerun()

    st.markdown("#### 2) URL del Google Sheet")
    url = st.text_input("Pega la URL de tu Google Sheet", value=cfg_get("gsheets_url", ""))
    if st.button("💾 Guardar URL") and url.strip():
        cfg_set("gsheets_url", url.strip())
        st.success("URL guardada."); st.rerun()

    st.markdown("#### 3) Sincronizar")
    if not (creds_ok and url_ok):
        st.info("Completa los pasos 1 y 2 para habilitar la sincronización.")
        return
    colA, colB = st.columns(2)
    with colA:
        if st.button("⬆️ Subir datos a Google Sheets"):
            try:
                n = sincronizar_a_sheets()
                st.success(f"Se subieron {n} tablas a Google Sheets.")
            except Exception as e:
                st.error(f"No se pudo sincronizar: {e}")
    with colB:
        if st.button("⬇️ Traer datos desde Google Sheets"):
            try:
                n = traer_de_sheets()
                st.success(f"Se trajeron {n} tablas desde Google Sheets."); st.rerun()
            except Exception as e:
                st.error(f"No se pudo traer: {e}")
    st.caption("Subir = tu app → Google Sheets. Traer = Google Sheets → tu app "
               "(reemplaza los datos locales).")


@st.dialog("✏️ Modificar usuario")
def dlg_modificar_usuario(usuarios, obras):
    um_sel = st.selectbox("Usuario a modificar", usuarios["nombre"].tolist(), key="dlg_us_sel")
    u = usuarios[usuarios["nombre"] == um_sel].iloc[0]
    e_codigo = st.text_input("Clave / código", value=u["codigo"] or "")
    e_nombre = st.text_input("Nombre completo", value=u["nombre"] or "")
    rol_idx = ROLES_ASIGNABLES.index(u["rol"]) if u["rol"] in ROLES_ASIGNABLES else 0
    e_rol = st.selectbox("Rol", ROLES_ASIGNABLES, index=rol_idx)
    e_tel = st.text_input("Teléfono", value=u["telefono"] or "")
    e_correo = st.text_input("Correo", value=u["correo"] or "")
    obra_op2 = ["(Todas / sin asignar)"] + (obras["nombre"].tolist() if not obras.empty else [])
    obra_actual = u["obra"] if pd.notna(u["obra"]) else "(Todas / sin asignar)"
    obra_idx = obra_op2.index(obra_actual) if obra_actual in obra_op2 else 0
    e_obra = st.selectbox("Obra asignada", obra_op2, index=obra_idx)
    st.caption("La contraseña de acceso se cambia en «Restablecer la clave».")
    c1, c2 = st.columns(2)
    if c1.button("💾 Guardar cambios", key="dlg_us_save") and e_nombre.strip():
        oid = obra_id_por_nombre(e_obra) if e_obra != "(Todas / sin asignar)" else None
        ejecutar("UPDATE usuarios SET codigo=?, nombre=?, rol=?, telefono=?, correo=?, "
                 "obra_id=? WHERE id=?",
                 (e_codigo.strip(), e_nombre.strip(), e_rol, e_tel, e_correo, oid, int(u["id"])))
        st.session_state["open_user_dlg"] = False
        st.rerun()
    if c2.button("Cancelar", key="dlg_us_cancel"):
        st.session_state["open_user_dlg"] = False
        st.rerun()


def vista_usuarios(rol: str):
    st.subheader("👤 Usuarios y roles")
    if not puede(rol, "usuarios"):
        st.warning("Solo el Administrador puede gestionar usuarios."); return
    usuarios = obtener_usuarios()
    if usuarios.empty:
        st.caption("Aún no hay usuarios. Da de alta el primero abajo.")
    else:
        vis = usuarios[["codigo", "nombre", "rol", "obra", "telefono", "correo"]].rename(
            columns={"codigo": "Clave", "nombre": "Nombre", "rol": "Rol",
                     "obra": "Obra asignada", "telefono": "Teléfono", "correo": "Correo"})
        st.dataframe(vis, width="stretch", hide_index=True)

    obras = obtener_obras()
    st.markdown("#### Dar de alta un usuario")
    st.caption("El usuario entrará a la plataforma con la clave de acceso que le asignes aquí.")
    with st.form("form_user", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            codigo = st.text_input("Clave / código de identificación (ej. SUP-01)")
            nombre = st.text_input("Nombre completo (ej. nombre del supervisor)")
            rol_u = st.selectbox("Rol", ROLES_ASIGNABLES)
            telefono = st.text_input("Teléfono")
            correo = st.text_input("Correo")
        with col2:
            obra_op = ["(Todas / sin asignar)"] + (obras["nombre"].tolist() if not obras.empty else [])
            obra_sel = st.selectbox("Obra asignada", obra_op)
            clave = st.text_input("Clave de acceso (contraseña)", type="password")
            clave2 = st.text_input("Repetir clave de acceso", type="password")
        crear = st.form_submit_button("➕ Crear usuario")
    if crear:
        if not nombre.strip():
            st.warning("Escribe el nombre del usuario.")
        elif not clave:
            st.warning("Asigna una clave de acceso.")
        elif clave != clave2:
            st.warning("Las claves no coinciden.")
        else:
            oid = obra_id_por_nombre(obra_sel) if obra_sel != "(Todas / sin asignar)" else None
            ejecutar("INSERT INTO usuarios(codigo,nombre,rol,telefono,correo,obra_id,clave_hash,activo) "
                     "VALUES(?,?,?,?,?,?,?,1)",
                     (codigo.strip(), nombre.strip(), rol_u, telefono, correo, oid, _hash(clave)))
            st.success(f"Usuario «{nombre}» creado con rol {rol_u}."); st.rerun()

    if not usuarios.empty:
        st.markdown("---")
        st.markdown("#### Modificar un usuario")
        st.caption("Solo el Administrador. Se pedirá su clave antes de abrir la ventana de edición.")
        if st.button("✏️ Modificar un usuario", key="btn_mod_user"):
            st.session_state["ask_user"] = True
            st.session_state["open_user_dlg"] = False
        if st.session_state.get("ask_user") and not st.session_state.get("open_user_dlg"):
            cl = st.text_input("Clave del Administrador", type="password", key="cl_user")
            cc1, cc2 = st.columns(2)
            if cc1.button("Abrir ventana de edición", key="ok_user"):
                if verificar_clave_admin(cl):
                    st.session_state["open_user_dlg"] = True
                    st.session_state["ask_user"] = False
                    st.rerun()
                else:
                    st.error("Clave del Administrador incorrecta.")
            if cc2.button("Cancelar", key="cancel_user_ask"):
                st.session_state["ask_user"] = False
                st.rerun()
        if st.session_state.get("open_user_dlg"):
            dlg_modificar_usuario(usuarios, obras)

    if not usuarios.empty:
        st.markdown("#### Restablecer la clave de un usuario")
        with st.form("form_reset_clave"):
            u_sel = st.selectbox("Usuario", usuarios["nombre"].tolist())
            nueva = st.text_input("Nueva clave", type="password")
            if st.form_submit_button("🔑 Cambiar clave") and nueva:
                ejecutar("UPDATE usuarios SET clave_hash=? WHERE nombre=?", (_hash(nueva), u_sel))
                st.success(f"Clave de «{u_sel}» actualizada."); st.rerun()

    st.markdown("---")
    st.markdown("#### 🔒 Seguridad del Administrador (recomendado al publicar en internet)")
    actual = cfg_get("admin_pass_hash")
    st.caption("Estado: " + ("✅ El Administrador ya requiere clave."
                             if actual else
                             "⚠️ El Administrador entra libre (bien para uso local, "
                             "riesgoso en internet)."))
    with st.form("form_admin_pass"):
        nueva_admin = st.text_input("Clave del Administrador (deja vacío y guarda para quitarla)",
                                    type="password")
        if st.form_submit_button("Guardar clave del Administrador"):
            if nueva_admin.strip():
                cfg_set("admin_pass_hash", _hash(nueva_admin.strip()))
                st.success("Clave del Administrador establecida.")
            else:
                cfg_set("admin_pass_hash", "")
                st.success("Clave del Administrador eliminada (acceso libre).")
            st.rerun()


def vista_proveedores(rol: str):
    st.subheader("🏭 Proveedores")
    if puede(rol, "todas_obras"):
        obra_prev = st.session_state.get("obra_trabajo_id")
        selector_obra_trabajo(obra_prev, rol, "🏢 Obra de trabajo (para cargar compras)",
                              "prov_obra_sel")
        st.caption("Los proveedores son compartidos entre todas las obras. Esta selección define "
                   "la obra a la que se cargarán las compras en la sección «Compras».")
    prov = obtener_proveedores()
    st.metric("Proveedores registrados", len(prov))
    if prov.empty:
        st.caption("Aún no hay proveedores. Registra el primero abajo.")
    else:
        st.markdown("#### Catálogo de proveedores")
        for _, p in prov.iterrows():
            cod = p["codigo"] if "codigo" in prov.columns and p["codigo"] else ""
            enc = f"🏭 {cod + ' · ' if cod else ''}{p['nombre']}  ·  Agente: {p['agente'] or '—'}"
            with st.expander(enc):
                st.write(f"📞 {p['telefono'] or '—'}  ·  ✉️ {p['correo'] or '—'}")
                st.write(f"**Clave:** {cod or '—'}")
                st.write(f"**Cuenta:** {p['cuenta'] or '—'}")
                st.write(f"**CLABE:** {p['clabe'] or '—'}")
                st.write(f"**Tarjeta:** {p['tarjeta'] or '—'}")
                banco_v = p["banco"] if "banco" in prov.columns and p["banco"] else "—"
                benef_v = p["beneficiario"] if "beneficiario" in prov.columns and p["beneficiario"] else "—"
                st.write(f"**Banco:** {banco_v}")
                st.write(f"**Beneficiario:** {benef_v}")
    if not puede(rol, "editar"):
        st.info("Solo Administrador o Ingeniero pueden registrar proveedores."); return
    st.markdown("---")
    st.markdown("#### Registrar nuevo proveedor")
    with st.form("form_prov", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            codigo = st.text_input("Clave del proveedor (identificador)")
            nombre = st.text_input("Proveedor (empresa)")
            agente = st.text_input("Nombre del agente de ventas")
            telefono = st.text_input("Número de contacto")
            correo = st.text_input("Correo")
        with col2:
            cuenta = st.text_input("Número de cuenta")
            clabe = st.text_input("CLABE interbancaria")
            tarjeta = st.text_input("Tarjeta")
            banco = st.text_input("Banco")
            beneficiario = st.text_input("Beneficiario de la cuenta")
        if st.form_submit_button("➕ Guardar proveedor") and nombre.strip():
            ejecutar("INSERT INTO proveedores(codigo,nombre,agente,telefono,correo,cuenta,clabe,"
                     "tarjeta,banco,beneficiario) VALUES(?,?,?,?,?,?,?,?,?,?)",
                     (codigo.strip(), nombre.strip(), agente, telefono, correo, cuenta, clabe,
                      tarjeta, banco, beneficiario))
            st.success(f"Proveedor «{nombre}» guardado."); st.rerun()


@st.dialog("✏️ Modificar obra")
def dlg_modificar_obra(obras, clientes):
    ob_labels, ob_map = opciones_clave(obras)
    ob_lbl = st.selectbox("Obra a modificar", ob_labels, key="dlg_ob_sel")
    o = obras[obras["nombre"] == ob_map[ob_lbl]].iloc[0]
    e_codigo = st.text_input("Clave de la obra", value=o["codigo"] or "")
    e_nombre = st.text_input("Nombre de la obra", value=o["nombre"] or "")
    cli_list = clientes["empresa"].tolist()
    cli_idx = cli_list.index(o["cliente"]) if o["cliente"] in cli_list else 0
    e_cliente = st.selectbox("Cliente", cli_list, index=cli_idx)
    e_ubic = st.text_input("Ubicación", value=o["ubicacion"] or "")
    e_ing = st.text_input("Responsable de obra", value=o["ingeniero"] or "")
    e_pres = st.number_input("Presupuesto ($ MXN)", min_value=0.0, step=10000.0,
                             value=float(o["presupuesto"] or 0), format="%.2f")
    est_op = ["En proceso", "Detenida", "Terminada"]
    e_estatus = st.selectbox("Estatus", est_op,
                             index=est_op.index(o["estatus"]) if o["estatus"] in est_op else 0)
    c1, c2 = st.columns(2)
    if c1.button("💾 Guardar cambios", key="dlg_ob_save") and e_nombre.strip():
        cid = int(clientes[clientes["empresa"] == e_cliente]["id"].iloc[0])
        ejecutar("UPDATE obras SET codigo=?, nombre=?, cliente_id=?, ubicacion=?, "
                 "ingeniero=?, presupuesto=?, estatus=? WHERE id=?",
                 (e_codigo.strip(), e_nombre.strip(), cid, e_ubic, e_ing, e_pres,
                  e_estatus, int(o["id"])))
        st.session_state["open_obra_dlg"] = False
        st.rerun()
    if c2.button("Cancelar", key="dlg_ob_cancel"):
        st.session_state["open_obra_dlg"] = False
        st.rerun()


@st.dialog("✏️ Modificar contratista")
def dlg_modificar_contratista(contr):
    cm_labels, cm_map = opciones_clave(contr)
    cm_lbl = st.selectbox("Contratista a modificar", cm_labels, key="dlg_co_sel")
    cc = contr[contr["nombre"] == cm_map[cm_lbl]].iloc[0]
    e_codigo = st.text_input("Clave del contratista", value=cc["codigo"] or "")
    e_nombre = st.text_input("Nombre del contratista", value=cc["nombre"] or "")
    e_esp = st.text_input("Especialidad", value=cc["especialidad"] or "")
    e_tel = st.text_input("Teléfono", value=cc["telefono"] or "")
    e_correo = st.text_input("Correo", value=cc["correo"] or "")
    e_monto = st.number_input("Monto contratado ($ MXN)", min_value=0.0, step=1000.0,
                              value=float(cc["monto_contratado"] or 0), format="%.2f",
                              help="Tope para el control de destajos de este contratista.")
    c1, c2 = st.columns(2)
    if c1.button("💾 Guardar cambios", key="dlg_co_save") and e_nombre.strip():
        ejecutar("UPDATE contratistas SET codigo=?, nombre=?, especialidad=?, telefono=?, "
                 "correo=?, monto_contratado=? WHERE id=?",
                 (e_codigo.strip(), e_nombre.strip(), e_esp, e_tel, e_correo, e_monto,
                  int(cc["id"])))
        st.session_state["open_contr_dlg"] = False
        st.rerun()
    if c2.button("Cancelar", key="dlg_co_cancel"):
        st.session_state["open_contr_dlg"] = False
        st.rerun()


def vista_obras(rol: str):
    st.subheader("🏢 Obras · Alta y listado")
    obras = obtener_obras()
    if not obras.empty:
        st.dataframe(obras[["codigo", "nombre", "cliente", "ubicacion", "ingeniero",
                            "presupuesto", "fecha_inicio", "fecha_fin", "estatus"]]
                     .rename(columns={"codigo": "Clave", "nombre": "Obra", "cliente": "Cliente",
                                      "ubicacion": "Ubicación", "ingeniero": "Responsable",
                                      "presupuesto": "Presupuesto", "fecha_inicio": "Inicio",
                                      "fecha_fin": "Término", "estatus": "Estatus"}),
                     width="stretch", hide_index=True)
    else:
        st.caption("Aún no hay obras registradas. Crea la primera abajo.")
    if not puede(rol, "admin"):
        st.info("Solo el Administrador puede dar de alta o modificar obras."); return

    # ----- Contrato aprobado (PDF) por obra -----
    if not obras.empty:
        st.markdown("#### 📄 Contrato aprobado (PDF)")
        ob_labels_c, ob_map_c = opciones_clave(obras)
        oc_lbl = st.selectbox("Selecciona la obra", ob_labels_c, key="contrato_obra_sel")
        oc_id = obra_id_por_nombre(ob_map_c[oc_lbl])
        actual = obtener_contrato(oc_id)
        if actual is not None:
            import base64
            st.success(f"Contrato cargado: {actual['nombre']}  ·  subido el {actual['fecha']}")
            st.download_button("📄 Descargar contrato aprobado",
                               data=base64.b64decode(actual["contenido"]),
                               file_name=actual["nombre"], mime="application/pdf",
                               key="dl_contrato")
        else:
            st.caption("Esta obra aún no tiene contrato aprobado cargado.")
        pdf_contrato = st.file_uploader("Cargar / reemplazar contrato aprobado (PDF)",
                                        type=["pdf"], key="up_contrato")
        if pdf_contrato is not None and st.button("💾 Guardar contrato", key="save_contrato"):
            guardar_contrato(oc_id, pdf_contrato)
            st.success("Contrato aprobado guardado."); st.rerun()

        link_actual = consultar("SELECT contrato_link FROM obras WHERE id=?",
                                (oc_id,))["contrato_link"].iloc[0]
        st.caption("Opción recomendada para internet: pega un link del contrato "
                   "(Google Drive, Dropbox, etc.). El link no se borra al reiniciarse el servidor.")
        nuevo_link = st.text_input("Link del contrato (opcional)", value=link_actual or "",
                                   key="contrato_link_in")
        if st.button("💾 Guardar link del contrato", key="save_link_contrato"):
            ejecutar("UPDATE obras SET contrato_link=? WHERE id=?", (nuevo_link.strip(), oc_id))
            st.success("Link del contrato guardado."); st.rerun()
        if link_actual:
            st.markdown(f"🔗 [Abrir contrato en el navegador]({link_actual})")
        st.markdown("---")

    st.markdown("#### Dar de alta una nueva obra")
    clientes = obtener_clientes()
    if clientes.empty:
        st.warning("Primero registra un cliente en el apartado «CRM».")
        return
    with st.form("form_obra", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            codigo = st.text_input("Clave de la obra (ej. TGZ01)")
            nombre = st.text_input("Nombre de la obra")
            cliente_sel = st.selectbox("Cliente", clientes["empresa"].tolist())
            ubicacion = st.text_input("Ubicación")
            usuarios = obtener_usuarios()
            if not usuarios.empty:
                ing_labels, ing_map = opciones_clave(usuarios)
                ing_lbl = st.selectbox("Responsable de obra", ing_labels)
                ingeniero = ing_map[ing_lbl]
            else:
                ingeniero = st.text_input("Responsable de obra (regístralo en «Usuarios»)")
        with col2:
            presupuesto = st.number_input("Presupuesto ($ MXN)", min_value=0.0, step=10000.0, format="%.2f")
            inicio = st.date_input("Inicio", HOY)
            fin = st.date_input("Término", HOY + timedelta(days=90))
            estatus = st.selectbox("Estatus", ["En proceso", "Detenida", "Terminada"])
        if st.form_submit_button("🏗️ Crear obra") and nombre.strip():
            cid = int(clientes[clientes["empresa"] == cliente_sel]["id"].iloc[0])
            ejecutar("INSERT INTO obras(codigo,nombre,cliente_id,ubicacion,ingeniero,presupuesto,"
                     "fecha_inicio,fecha_fin,estatus) VALUES(?,?,?,?,?,?,?,?,?)",
                     (codigo.strip(), nombre.strip(), cid, ubicacion, ingeniero, presupuesto,
                      inicio.isoformat(), fin.isoformat(), estatus))
            oid = obra_id_por_nombre(nombre.strip())
            for et, i, f in [("Cimentación", inicio, inicio + timedelta(days=20)),
                             ("Estructura", inicio + timedelta(days=15), inicio + timedelta(days=50)),
                             ("Acabados", inicio + timedelta(days=45), fin)]:
                ejecutar("INSERT INTO etapas(obra_id,etapa,inicio,fin,estado,avance) "
                         "VALUES(?,?,?,?,?,?)",
                         (oid, et, i.isoformat(), f.isoformat(), "Por iniciar", 0))
            st.success(f"Obra «{nombre}» creada con sus etapas base."); st.rerun()

    # ----- Modificar una obra (botón -> clave -> ventana) -----
    if not obras.empty:
        st.markdown("---")
        st.markdown("#### Modificar una obra")
        st.caption("Solo el Administrador. Se pedirá su clave antes de abrir la ventana de edición.")
        if st.button("✏️ Modificar una obra", key="btn_mod_obra"):
            st.session_state["ask_obra"] = True
            st.session_state["open_obra_dlg"] = False
        if st.session_state.get("ask_obra") and not st.session_state.get("open_obra_dlg"):
            cl = st.text_input("Clave del Administrador", type="password", key="cl_obra")
            cc1, cc2 = st.columns(2)
            if cc1.button("Abrir ventana de edición", key="ok_obra"):
                if verificar_clave_admin(cl):
                    st.session_state["open_obra_dlg"] = True
                    st.session_state["ask_obra"] = False
                    st.rerun()
                else:
                    st.error("Clave del Administrador incorrecta.")
            if cc2.button("Cancelar", key="cancel_obra_ask"):
                st.session_state["ask_obra"] = False
                st.rerun()
        if st.session_state.get("open_obra_dlg"):
            dlg_modificar_obra(obras, clientes)


def vista_contratistas(rol: str):
    st.subheader("👷 Contratistas · Alta y asignación a obra")
    contr = consultar("SELECT * FROM contratistas ORDER BY nombre")
    obras = obtener_obras()
    st.markdown("#### Contratistas registrados y sus obras")
    if contr.empty:
        st.caption("Aún no hay contratistas registrados.")
    else:
        for _, c in contr.iterrows():
            cod = c["codigo"] if "codigo" in contr.columns and c["codigo"] else ""
            enc = f"👷 {cod + ' · ' if cod else ''}{c['nombre']}  ·  {c['especialidad']}  ·  📞 {c['telefono']}"
            asign = consultar("SELECT o.nombre AS obra, d.concepto, d.monto_contratado, d.estatus "
                              "FROM destajos d JOIN obras o ON d.obra_id=o.id "
                              "WHERE d.contratista=?", (c["nombre"],))
            with st.expander(enc):
                ctrl = control_contratista(c["nombre"])
                m1, m2, m3 = st.columns(3)
                m1.metric("Monto contratado", pesos(ctrl["cap"]))
                m2.metric("Asignado en destajos", pesos(ctrl["asignado"]))
                m3.metric("Pagado", pesos(ctrl["pagado"]))
                if ctrl["excedido"]:
                    st.error("⚠️ CONTRATO EXCEDIDO")
                elif ctrl["cap"] > 0:
                    st.caption(f"Disponible por asignar: {pesos(ctrl['disponible'])}")
                else:
                    st.caption("Sin monto contratado definido (sin control de tope).")
                if asign.empty:
                    st.caption("Sin obras asignadas todavía.")
                else:
                    t = asign.copy()
                    t["monto_contratado"] = t["monto_contratado"].map(lambda x: f"{pesos(x)}")
                    st.dataframe(t.rename(columns={"obra": "Obra", "concepto": "Concepto",
                                 "monto_contratado": "Contratado", "estatus": "Estatus"}),
                                 width="stretch", hide_index=True)
    if not puede(rol, "editar"):
        st.info("Solo Administrador o Ingeniero pueden registrar/asignar contratistas."); return
    st.markdown("---")
    st.markdown("#### 1) Registrar un contratista nuevo")
    with st.form("form_contr", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            codigo = st.text_input("Clave del contratista (identificador)")
            nombre = st.text_input("Nombre del contratista")
            especialidad = st.text_input("Especialidad (ej. Albañilería)")
        with col2:
            telefono = st.text_input("Teléfono"); correo = st.text_input("Correo")
            monto_contr = st.number_input("Monto contratado ($ MXN)", min_value=0.0, step=1000.0,
                                          format="%.2f",
                                          help="Tope para el control de destajos de este contratista.")
        if st.form_submit_button("➕ Registrar contratista") and nombre.strip():
            ejecutar("INSERT INTO contratistas(codigo,nombre,especialidad,telefono,correo,"
                     "monto_contratado) VALUES(?,?,?,?,?,?)",
                     (codigo.strip(), nombre.strip(), especialidad, telefono, correo, monto_contr))
            st.success(f"Contratista «{nombre}» registrado."); st.rerun()
    st.markdown("#### 2) Asignar un contratista a una obra")
    if contr.empty or obras.empty:
        st.caption("Necesitas al menos un contratista y una obra registrados.")
    else:
        c_labels, c_map = opciones_clave(contr)
        o_labels, o_map = opciones_clave(obras)
        with st.form("form_asignar", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                contr_lbl = st.selectbox("Contratista", c_labels)
                obra_lbl = st.selectbox("Obra a asignar", o_labels)
                concepto = st.text_input("Concepto del trabajo")
            with col2:
                monto = st.number_input("Monto contratado ($ MXN)", min_value=0.0, step=1000.0, format="%.2f")
                anticipo = st.number_input("Anticipo / pagado ($ MXN)", min_value=0.0, step=1000.0, format="%.2f")
                avance = st.slider("% Avance", 0, 100, 0)
            if st.form_submit_button("🔗 Asignar a la obra") and concepto.strip():
                oid = obra_id_por_nombre(o_map[obra_lbl])
                ejecutar("INSERT INTO destajos(obra_id,contratista,concepto,monto_contratado,"
                         "pagado,avance,estatus) VALUES(?,?,?,?,?,?,?)",
                         (oid, c_map[contr_lbl], concepto.strip(), monto, anticipo, avance, "En proceso"))
                st.success(f"«{c_map[contr_lbl]}» asignado a «{o_map[obra_lbl]}»."); st.rerun()

    # ----- Contrato aprobado (PDF) del contratista -----
    if not contr.empty:
        st.markdown("---")
        st.markdown("#### 📄 Contrato aprobado del contratista (PDF)")
        cc_labels, cc_map = opciones_clave(contr)
        cc_lbl = st.selectbox("Selecciona el contratista", cc_labels, key="contrato_contr_sel")
        cc_row = contr[contr["nombre"] == cc_map[cc_lbl]].iloc[0]
        cc_id = int(cc_row["id"])
        actual_c = obtener_contrato_contr(cc_id)
        if actual_c is not None:
            import base64
            st.success(f"Contrato cargado: {actual_c['nombre']}  ·  subido el {actual_c['fecha']}")
            st.download_button("📄 Descargar contrato del contratista",
                               data=base64.b64decode(actual_c["contenido"]),
                               file_name=actual_c["nombre"], mime="application/pdf",
                               key="dl_contrato_contr")
        else:
            st.caption("Este contratista aún no tiene contrato aprobado cargado.")
        pdf_cc = st.file_uploader("Cargar / reemplazar contrato aprobado (PDF)",
                                  type=["pdf"], key="up_contrato_contr")
        if pdf_cc is not None and st.button("💾 Guardar contrato", key="save_contrato_contr"):
            guardar_contrato_contr(cc_id, pdf_cc)
            st.success("Contrato aprobado guardado."); st.rerun()
        link_cc = cc_row["contrato_link"] if "contrato_link" in contr.columns else None
        st.caption("Opción recomendada para internet: pega un link del contrato "
                   "(Google Drive, Dropbox, etc.). El link no se borra al reiniciarse el servidor.")
        nuevo_link_cc = st.text_input("Link del contrato (opcional)", value=link_cc or "",
                                      key="contrato_contr_link_in")
        if st.button("💾 Guardar link del contrato", key="save_link_contr"):
            ejecutar("UPDATE contratistas SET contrato_link=? WHERE id=?",
                     (nuevo_link_cc.strip(), cc_id))
            st.success("Link del contrato guardado."); st.rerun()
        if link_cc:
            st.markdown(f"🔗 [Abrir contrato en el navegador]({link_cc})")

    # ----- Modificar un contratista (botón -> clave -> ventana) -----
    if puede(rol, "admin") and not contr.empty:
        st.markdown("---")
        st.markdown("#### Modificar un contratista")
        st.caption("Solo el Administrador. Se pedirá su clave antes de abrir la ventana de edición.")
        if st.button("✏️ Modificar un contratista", key="btn_mod_contr"):
            st.session_state["ask_contr"] = True
            st.session_state["open_contr_dlg"] = False
        if st.session_state.get("ask_contr") and not st.session_state.get("open_contr_dlg"):
            cl = st.text_input("Clave del Administrador", type="password", key="cl_contr")
            cc1, cc2 = st.columns(2)
            if cc1.button("Abrir ventana de edición", key="ok_contr"):
                if verificar_clave_admin(cl):
                    st.session_state["open_contr_dlg"] = True
                    st.session_state["ask_contr"] = False
                    st.rerun()
                else:
                    st.error("Clave del Administrador incorrecta.")
            if cc2.button("Cancelar", key="cancel_contr_ask"):
                st.session_state["ask_contr"] = False
                st.rerun()
        if st.session_state.get("open_contr_dlg"):
            dlg_modificar_contratista(contr)


def vista_presupuesto(obra_id: int, rol: str):
    st.subheader("💵 Presupuesto de obra")
    if not requiere_obra(obra_id):
        return
    pres = consultar("SELECT * FROM presupuesto WHERE obra_id=? ORDER BY partida", (obra_id,))
    k = calcular_kpis(obra_id)
    total_cargado = pres["monto"].sum() if not pres.empty else 0.0
    c1, c2, c3 = st.columns(3)
    c1.metric("Presupuesto cargado", f"{pesos(total_cargado)}")
    c2.metric("Presupuesto de la obra", f"{pesos(k['presupuesto'])}")
    c3.metric("Ejercido (compras)", f"{pesos(k['ejercido'])}", f"{k['pct']}%")
    if not pres.empty:
        st.plotly_chart(grafica_presupuesto_partidas(pres), width="stretch", key="plt_7")
        comp = pd.DataFrame({"Concepto": ["Presupuesto cargado", "Ejercido"],
                             "Monto": [total_cargado, k["ejercido"]]})
        comp["_etq"] = comp["Monto"].map(pesos)
        fig = px.bar(comp, x="Concepto", y="Monto", text="_etq", color="Concepto",
                     color_discrete_sequence=[COLOR_PRIMARIO, COLOR_ACENTO],
                     title="Presupuesto cargado vs Ejercido")
        fig.update_traces(textposition="outside")
        st.plotly_chart(_mate(fig, 340), width="stretch", key="plt_8")
        tabla = pres[["partida", "concepto", "monto"]].copy()
        tabla["monto"] = tabla["monto"].map(lambda x: f"{pesos(x)}")
        st.dataframe(tabla.rename(columns={"partida": "Partida", "concepto": "Concepto",
                     "monto": "Monto"}), width="stretch", hide_index=True)
    else:
        st.caption("Aún no se ha cargado presupuesto para esta obra.")
    if not puede(rol, "editar"):
        st.info("Solo Administrador o Ingeniero pueden cargar presupuesto."); return
    st.markdown("---")
    st.markdown("#### Cargar una partida (manual)")
    with st.form("form_pres", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        with col1: partida = st.text_input("Partida (ej. Albañilería)")
        with col2: concepto = st.text_input("Concepto")
        with col3: monto = st.number_input("Monto ($ MXN)", min_value=0.0, step=1000.0, format="%.2f")
        if st.form_submit_button("➕ Agregar partida") and partida.strip():
            ejecutar("INSERT INTO presupuesto(obra_id,partida,concepto,monto) VALUES(?,?,?,?)",
                     (obra_id, partida.strip(), concepto, monto))
            st.success("Partida agregada al presupuesto."); st.rerun()
    st.markdown("#### 📄 Cargar el presupuesto aprobado (PDF)")
    actual_pdf = obtener_presupuesto_pdf(obra_id)
    if actual_pdf is not None:
        import base64
        st.success(f"Presupuesto cargado: {actual_pdf['nombre']}  ·  subido el {actual_pdf['fecha']}")
        st.download_button("📄 Descargar presupuesto aprobado",
                           data=base64.b64decode(actual_pdf["contenido"]),
                           file_name=actual_pdf["nombre"], mime="application/pdf",
                           key="dl_pres_pdf")
    else:
        st.caption("Esta obra aún no tiene presupuesto aprobado cargado.")
    pdf_pres = st.file_uploader("Cargar / reemplazar presupuesto aprobado (PDF)",
                                type=["pdf"], key="up_pres_pdf")
    if pdf_pres is not None and st.button("💾 Guardar presupuesto (PDF)", key="save_pres_pdf"):
        guardar_presupuesto_pdf(obra_id, pdf_pres)
        st.success("Presupuesto aprobado guardado."); st.rerun()

    link_pres = consultar("SELECT presupuesto_link FROM obras WHERE id=?",
                          (obra_id,))["presupuesto_link"].iloc[0]
    st.caption("Opción recomendada para internet: pega un link del presupuesto "
               "(Google Drive, Dropbox, etc.). El link no se borra al reiniciarse el servidor.")
    nuevo_link = st.text_input("Link del presupuesto (opcional)", value=link_pres or "",
                               key="pres_link_in")
    if st.button("💾 Guardar link del presupuesto", key="save_pres_link"):
        ejecutar("UPDATE obras SET presupuesto_link=? WHERE id=?", (nuevo_link.strip(), obra_id))
        st.success("Link del presupuesto guardado."); st.rerun()
    if link_pres:
        st.markdown(f"🔗 [Abrir presupuesto en el navegador]({link_pres})")


def vista_control_financiero(obra_id: int, rol: str):
    st.subheader("💰 Control financiero de la obra")
    if not requiere_obra(obra_id):
        return
    abonos = consultar("SELECT * FROM abonos WHERE obra_id=? ORDER BY fecha DESC", (obra_id,))
    k = calcular_kpis(obra_id)
    recibido = total_abonos(obra_id)
    egresos = k["ejercido"]
    balance = recibido - egresos
    c1, c2, c3 = st.columns(3)
    c1.metric("Total recibido (abonos)", f"{pesos(recibido)}")
    c2.metric("Egresos (compras)", f"{pesos(egresos)}")
    c3.metric("Balance", f"{pesos(balance)}",
              "A favor" if balance >= 0 else "En contra")
    if not abonos.empty:
        tabla = abonos[["fecha", "concepto", "metodo_pago", "monto", "nota"]].copy()
        tabla["monto"] = tabla["monto"].map(lambda x: f"{pesos(x)}")
        st.dataframe(tabla.rename(columns={"fecha": "Fecha", "concepto": "Concepto",
                     "metodo_pago": "Método de pago", "monto": "Monto", "nota": "Nota"}),
                     width="stretch", hide_index=True)
    else:
        st.caption("Aún no hay abonos o pagos registrados para esta obra.")
    if not puede(rol, "editar"):
        st.info("Tu rol es de solo lectura."); return
    st.markdown("---")
    st.markdown("#### ➕ Registrar un pago o abono recibido")
    with st.form("form_abono", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            fecha = st.date_input("Fecha del abono", HOY)
            metodo = st.selectbox("Método de pago", METODOS_PAGO)
        with col2:
            concepto = st.text_input("Concepto (ej. Anticipo, Estimación 1)")
            monto = st.number_input("Monto recibido ($ MXN)", min_value=0.0, step=1000.0,
                                    format="%.2f")
        with col3:
            nota = st.text_area("Nota (opcional)", height=90)
        if st.form_submit_button("💾 Registrar abono") and concepto.strip() and monto > 0:
            ejecutar("INSERT INTO abonos(obra_id,fecha,concepto,monto,metodo_pago,nota) "
                     "VALUES(?,?,?,?,?,?)",
                     (obra_id, fecha.isoformat(), concepto.strip(), monto, metodo, nota.strip()))
            st.success(f"Abono de {pesos(monto)} registrado."); st.rerun()
    if not abonos.empty:
        st.markdown("#### 🗑️ Eliminar un abono")
        op = {f"{r['fecha']} · {r['concepto']} · {pesos(r['monto'])}": int(r["id"])
              for _, r in abonos.iterrows()}
        sel = st.selectbox("Selecciona el abono a eliminar", list(op.keys()), key="del_abono_sel")
        if st.button("Eliminar abono seleccionado", key="del_abono_btn"):
            ejecutar("DELETE FROM abonos WHERE id=?", (op[sel],))
            st.success("Abono eliminado."); st.rerun()


def vista_requisiciones(obra_id: int, rol: str, usuario: str):
    st.subheader("📦 Requisiciones de materiales")
    if not requiere_obra(obra_id):
        return
    reqs = consultar("SELECT * FROM requisiciones WHERE obra_id=? ORDER BY fecha DESC", (obra_id,))
    if not reqs.empty:
        c1, c2, c3 = st.columns(3)
        c1.metric("Requisiciones", len(reqs))
        c2.metric("Pendientes", int((reqs["estatus"] == "Solicitada").sum()))
        c3.metric("Costo estimado", f"{pesos(reqs['costo_estimado'].sum())}")
        tabla = reqs[["folio", "fecha", "material", "cantidad", "unidad", "proveedor",
                      "costo_estimado", "estatus"]].copy()
        tabla["costo_estimado"] = tabla["costo_estimado"].map(lambda x: f"{pesos(x)}")
        st.dataframe(tabla.rename(columns={"folio": "Folio", "fecha": "Fecha",
                     "material": "Material", "cantidad": "Cant.", "unidad": "Unidad",
                     "proveedor": "Proveedor", "costo_estimado": "Costo est.",
                     "estatus": "Estatus"}), width="stretch", hide_index=True)
    else:
        st.caption("No hay requisiciones para esta obra todavía.")
    if not puede(rol, "editar"):
        st.info("Tu rol es de solo lectura."); return
    st.markdown("#### Nueva requisición")
    cat_prov = obtener_proveedores()
    rp_labels, rp_map = opciones_clave(cat_prov) if not cat_prov.empty else ([], {})
    with st.form("form_req", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            folio = st.text_input("Folio", value=f"REQ-{len(reqs)+1:03d}")
            material = st.text_input("Material")
        with col2:
            cantidad = st.number_input("Cantidad", min_value=0.0, step=1.0)
            unidad = st.text_input("Unidad", value="pza")
        with col3:
            if rp_labels:
                rp_lbl = st.selectbox("Proveedor (del catálogo)", rp_labels)
                proveedor = rp_map[rp_lbl]
            else:
                proveedor = st.text_input("Proveedor (regístralo en «Proveedores»)")
            costo = st.number_input("Costo estimado ($ MXN)", min_value=0.0, step=100.0, format="%.2f")
            prioridad = st.selectbox("Prioridad", PRIORIDAD_REQ)
        if st.form_submit_button("➕ Registrar requisición") and material.strip():
            ejecutar("INSERT INTO requisiciones(obra_id,folio,fecha,solicitante,material,cantidad,"
                     "unidad,proveedor,costo_estimado,estatus,prioridad) "
                     "VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                     (obra_id, folio, HOY.isoformat(), usuario, material.strip(), cantidad,
                      unidad, proveedor, costo, "Solicitada", prioridad))
            st.success("Requisición registrada."); st.rerun()
    if not reqs.empty:
        st.markdown("#### Actualizar estatus / prioridad de una requisición")
        with st.form("form_req_estatus"):
            folio_sel = st.selectbox("Folio", reqs["folio"].tolist())
            cse1, cse2 = st.columns(2)
            with cse1:
                nuevo = st.selectbox("Nuevo estatus", ESTATUS_REQ)
            with cse2:
                nueva_prio = st.selectbox("Prioridad", PRIORIDAD_REQ)
            if st.form_submit_button("💾 Actualizar"):
                ejecutar("UPDATE requisiciones SET estatus=?, prioridad=? WHERE obra_id=? AND folio=?",
                         (nuevo, nueva_prio, obra_id, folio_sel))
                st.success(f"Requisición {folio_sel} → {nuevo} · {nueva_prio}."); st.rerun()

        st.markdown("#### 📤 Enviar requisición (PDF / WhatsApp / correo)")
        folio_wa = st.selectbox("Requisición a enviar", reqs["folio"].tolist(), key="wa_folio")
        row = reqs[reqs["folio"] == folio_wa].iloc[0]
        rid = int(row["id"])
        if FPDF_OK:
            st.download_button("⬇️ Descargar PDF", data=pdf_requisicion(rid),
                               file_name=f"{folio_wa}.pdf", mime="application/pdf", key="reqpdf")
            resumen = (f"Requisición {row['folio']} - Obra: "
                       f"{consultar('SELECT nombre FROM obras WHERE id=?', (obra_id,))['nombre'].iloc[0]}\n"
                       f"Material: {row['material']} ({row['cantidad']} {row['unidad']})\n"
                       f"Proveedor: {row['proveedor'] or 'N/D'}\n"
                       f"Costo estimado: {pesos(row['costo_estimado'])}\n"
                       f"Estatus: {row['estatus']}")
            bloque_enviar_reporte(pdf_requisicion(rid), f"Requisicion {row['folio']}",
                                  f"{folio_wa}.pdf", resumen, "req")
        else:
            st.caption("Para el PDF instala una vez: python -m pip install fpdf2")


def vista_destajos(obra_id: int, rol: str):
    st.subheader("🔨 Destajos de contratistas")
    if not requiere_obra(obra_id):
        return
    dest = consultar("SELECT * FROM destajos WHERE obra_id=? ORDER BY contratista", (obra_id,))
    if not dest.empty:
        dest["saldo"] = dest["monto_contratado"] - dest["pagado"]
        c1, c2, c3 = st.columns(3)
        c1.metric("Contratado", f"{pesos(dest['monto_contratado'].sum())}")
        c2.metric("Pagado", f"{pesos(dest['pagado'].sum())}")
        c3.metric("Saldo por pagar", f"{pesos(dest['saldo'].sum())}")
        st.plotly_chart(grafica_destajos(dest), width="stretch", key="plt_9")
        tabla = dest[["contratista", "concepto", "monto_contratado", "pagado", "saldo",
                      "avance", "estatus"]].copy()
        for col in ["monto_contratado", "pagado", "saldo"]:
            tabla[col] = tabla[col].map(lambda x: f"{pesos(x)}")
        st.dataframe(tabla.rename(columns={"contratista": "Contratista", "concepto": "Concepto",
                     "monto_contratado": "Contratado", "pagado": "Pagado", "saldo": "Saldo",
                     "avance": "% Avance", "estatus": "Estatus"}), width="stretch", hide_index=True)

        # Control del monto contratado por contratista (tope global)
        st.markdown("#### 🔒 Control de monto contratado")
        filas, excedidos = [], []
        for nom in dest["contratista"].unique():
            ct = control_contratista(nom)
            estado = ("⚠️ CONTRATO EXCEDIDO" if ct["excedido"]
                      else ("Dentro del contrato" if ct["cap"] > 0 else "Sin tope definido"))
            filas.append({"Contratista": nom, "Monto contratado": pesos(ct["cap"]),
                          "Asignado en destajos": pesos(ct["asignado"]),
                          "Pagado": pesos(ct["pagado"]), "Estado": estado})
            if ct["excedido"]:
                excedidos.append((nom, ct))
        st.dataframe(pd.DataFrame(filas), width="stretch", hide_index=True)
        for nom, ct in excedidos:
            st.error(f"⚠️ CONTRATO EXCEDIDO — {nom}: el monto contratado es "
                     f"{pesos(ct['cap'])} y ya lleva {pesos(max(ct['asignado'], ct['pagado']))}.")

        st.markdown("#### 📄 Reporte semanal de pagos")
        st.caption("Suma de todos los destajos por pagar de esta obra, con datos de la empresa y logo.")
        if FPDF_OK:
            st.download_button("📄 Descargar reporte semanal (PDF)",
                               data=pdf_pagos_semana(obra_id),
                               file_name=f"Pagos_semana_{HOY.isoformat()}.pdf",
                               mime="application/pdf", key="pdf_pagos_sem")
            resumen_p = (f"Reporte semanal de pagos a destajos\n"
                         f"Obra activa - {EMPRESA} - {HOY.isoformat()}")
            bloque_enviar_reporte(pdf_pagos_semana(obra_id), "Reporte semanal de pagos",
                                  f"Pagos_semana_{HOY.isoformat()}.pdf", resumen_p, "pagos")
        else:
            st.caption("Para el PDF instala una vez: python -m pip install fpdf2")
    else:
        st.caption("No hay destajos para esta obra todavía.")
    if not puede(rol, "editar"):
        st.info("Tu rol es de solo lectura."); return
    st.markdown("#### Nuevo destajo / contrato")
    cat_contr = consultar("SELECT * FROM contratistas ORDER BY nombre")
    c_labels, c_map = opciones_clave(cat_contr) if not cat_contr.empty else ([], {})
    with st.form("form_dest", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            if c_labels:
                c_lbl = st.selectbox("Contratista (del catálogo)", c_labels)
                contratista = c_map[c_lbl]
            else:
                contratista = st.text_input("Contratista (regístralo en «Contratistas»)")
            concepto = st.text_input("Concepto")
            estatus = st.selectbox("Estatus", ["En proceso", "Terminado", "Detenido"])
        with col2:
            monto = st.number_input("Monto contratado ($ MXN)", min_value=0.0, step=1000.0, format="%.2f")
            pagado = st.number_input("Pagado a la fecha ($ MXN)", min_value=0.0, step=1000.0, format="%.2f")
            avance = st.slider("% Avance", 0, 100, 0)
        if st.form_submit_button("➕ Registrar destajo") and (contratista or "").strip():
            ejecutar("INSERT INTO destajos(obra_id,contratista,concepto,monto_contratado,pagado,"
                     "avance,estatus) VALUES(?,?,?,?,?,?,?)",
                     (obra_id, contratista.strip(), concepto, monto, pagado, avance, estatus))
            st.success("Destajo registrado."); st.rerun()
    if not dest.empty:
        st.markdown("#### Registrar un pago a destajo")
        with st.form("form_dest_pago"):
            contr_sel = st.selectbox("Contratista", dest["contratista"].tolist())
            col1, col2 = st.columns(2)
            with col1:
                abono = st.number_input("Monto del pago ($ MXN)", min_value=0.0, step=1000.0,
                                        format="%.2f")
                metodo_d = st.selectbox("Método de pago", METODOS_PAGO)
            with col2:
                datos_banc = st.text_input("No. tarjeta / CLABE interbancaria / No. cuenta")
                banco_benef = st.text_input("Banco y nombre del beneficiario")
            if st.form_submit_button("💾 Aplicar pago"):
                actual = consultar("SELECT pagado FROM destajos WHERE obra_id=? AND contratista=?",
                                   (obra_id, contr_sel))["pagado"].iloc[0]
                ejecutar("UPDATE destajos SET pagado=?, metodo_pago=?, datos_bancarios=?, "
                         "banco_beneficiario=? WHERE obra_id=? AND contratista=?",
                         (float(actual) + abono, metodo_d, datos_banc.strip(),
                          banco_benef.strip(), obra_id, contr_sel))
                st.success(f"Pago de {pesos(abono)} aplicado a {contr_sel}."); st.rerun()


def vista_avances(obra_id: int, rol: str, usuario: str):
    st.subheader("📝 Avances y bitácora")
    if not requiere_obra(obra_id):
        return
    if not puede(rol, "editar"):
        st.info("Tu rol (Cliente) es de solo lectura.")
    etapas = consultar("SELECT * FROM etapas WHERE obra_id=?", (obra_id,))
    if puede(rol, "editar") and not etapas.empty:
        st.markdown("#### Actualizar avance de una etapa")
        with st.form("form_avance"):
            etapa_sel = st.selectbox("Etapa", etapas["etapa"].tolist())
            fila = etapas[etapas["etapa"] == etapa_sel].iloc[0]
            nuevo = st.slider("% de avance", 0, 100, int(fila["avance"]))
            estado = st.selectbox("Estado", ["Por iniciar", "En proceso", "Completada"],
                                  index=["Por iniciar", "En proceso", "Completada"].index(fila["estado"]))
            if st.form_submit_button("💾 Guardar avance"):
                ejecutar("UPDATE etapas SET avance=?, estado=? WHERE id=?",
                         (nuevo, estado, int(fila["id"])))
                st.success(f"Etapa «{etapa_sel}» → {nuevo}%."); st.rerun()
        st.markdown("---")
    if puede(rol, "editar"):
        st.markdown("#### Nueva nota de bitácora")
        with st.form("form_bit", clear_on_submit=True):
            fecha = st.date_input("Fecha", HOY)
            nota = st.text_area("Observaciones", height=100)
            if st.form_submit_button("➕ Registrar nota") and nota.strip():
                ejecutar("INSERT INTO bitacora(obra_id,fecha,autor,nota) VALUES(?,?,?,?)",
                         (obra_id, fecha.isoformat(), usuario, nota.strip()))
                st.success("Nota registrada."); st.rerun()
    st.markdown("#### Historial de bitácora")
    bit = consultar("SELECT fecha,autor,nota FROM bitacora WHERE obra_id=? ORDER BY fecha DESC",
                    (obra_id,))
    if bit.empty:
        st.caption("Sin notas registradas.")
    else:
        st.dataframe(bit.rename(columns={"fecha": "Fecha", "autor": "Autor", "nota": "Nota"}),
                     width="stretch", hide_index=True)


# =============================================================================
# 8B) EDITAR / BORRAR
# =============================================================================
def _val(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, np.integer):
        return int(v)
    if isinstance(v, np.floating):
        return float(v)
    if isinstance(v, (pd.Timestamp, datetime, date)):
        return pd.to_datetime(v).strftime("%Y-%m-%d")
    return v


def aplicar_cambios(tabla, df_orig, df_edit, columnas, defaults=None) -> int:
    defaults = defaults or {}
    cambios = 0
    orig_ids = set(int(i) for i in df_orig["id"].dropna().tolist())
    edit_ids = set()
    for _, row in df_edit.iterrows():
        rid = row.get("id")
        if pd.notna(rid):
            rid = int(rid); edit_ids.add(rid)
            sets = ", ".join(f"{c}=?" for c in columnas)
            vals = [_val(row.get(c)) for c in columnas] + [rid]
            ejecutar(f"UPDATE {tabla} SET {sets} WHERE id=?", tuple(vals)); cambios += 1
        else:
            if _val(row.get(columnas[0])) in (None, ""):
                continue
            cols = list(columnas) + list(defaults.keys())
            ph = ",".join(["?"] * len(cols))
            vals = [_val(row.get(c)) for c in columnas] + list(defaults.values())
            ejecutar(f"INSERT INTO {tabla}({','.join(cols)}) VALUES({ph})", tuple(vals)); cambios += 1
    for rid in orig_ids - edit_ids:
        if tabla == "obras":
            for hijo in ["etapas", "compras", "requisiciones", "destajos", "bitacora", "presupuesto"]:
                ejecutar(f"DELETE FROM {hijo} WHERE obra_id=?", (rid,))
        ejecutar(f"DELETE FROM {tabla} WHERE id=?", (rid,)); cambios += 1
    return cambios


def editor_tabla(clave, tabla, df, columnas, defaults=None):
    edit = st.data_editor(df, num_rows="dynamic", hide_index=True, width="stretch",
                          key=f"ed_{clave}", disabled=["id"], column_config={"id": None})
    if st.button("💾 Guardar cambios", key=f"btn_{clave}"):
        n = aplicar_cambios(tabla, df, edit, columnas, defaults)
        st.success(f"Cambios guardados ({n} registro(s) afectado(s))."); st.rerun()


def vista_editar(obra_id: int, rol: str, usuario: str):
    st.subheader("✏️ Editar / Borrar registros")
    if not puede(rol, "editar"):
        st.warning("Requiere permisos de Administrador o Ingeniero."); return
    st.info("Edita una celda y presiona Guardar. Para borrar un renglón, selecciónalo y "
            "usa la tecla de borrar; luego Guarda. Para agregar, escribe en el último renglón.")
    tabs = st.tabs(["Obras", "Clientes", "Proveedores", "Contratistas", "Compras",
                    "Presupuesto", "Destajos", "Requisiciones", "Etapas", "Bitácora", "Usuarios"])
    with tabs[0]:
        if not puede(rol, "admin"):
            st.caption("Solo el Administrador puede editar o borrar obras.")
        else:
            st.markdown("##### Obras  ⚠️ *borrar una obra elimina también sus datos asociados*")
            df = consultar("SELECT id,codigo,nombre,ubicacion,ingeniero,presupuesto,fecha_inicio,"
                           "fecha_fin,estatus FROM obras ORDER BY id")
            editor_tabla("obras", "obras", df, ["codigo", "nombre", "ubicacion", "ingeniero",
                         "presupuesto", "fecha_inicio", "fecha_fin", "estatus"])
    with tabs[1]:
        if not puede(rol, "crm"):
            st.caption("Solo el Administrador puede editar clientes.")
        else:
            df = consultar("SELECT id,empresa,contacto,telefono,correo,tipo,notas FROM clientes ORDER BY id")
            editor_tabla("clientes", "clientes", df,
                         ["empresa", "contacto", "telefono", "correo", "tipo", "notas"])
    with tabs[2]:
        df = consultar("SELECT id,codigo,nombre,agente,telefono,correo,cuenta,clabe,tarjeta "
                       "FROM proveedores ORDER BY id")
        editor_tabla("proveedores", "proveedores", df,
                     ["codigo", "nombre", "agente", "telefono", "correo", "cuenta", "clabe", "tarjeta"])
    with tabs[3]:
        df = consultar("SELECT id,codigo,nombre,especialidad,telefono,correo FROM contratistas ORDER BY id")
        editor_tabla("contratistas", "contratistas", df,
                     ["codigo", "nombre", "especialidad", "telefono", "correo"])
    with tabs[4]:
        if obra_id == -1:
            st.caption("Selecciona una obra activa para editar sus compras.")
        else:
            df = consultar("SELECT id,fecha,categoria,descripcion,importe,proveedor,comprador "
                           "FROM compras WHERE obra_id=? ORDER BY id", (obra_id,))
            editor_tabla("compras", "compras", df, ["fecha", "categoria", "descripcion",
                         "importe", "proveedor", "comprador"], defaults={"obra_id": obra_id})
    with tabs[5]:
        if obra_id == -1:
            st.caption("Selecciona una obra activa.")
        else:
            df = consultar("SELECT id,partida,concepto,monto FROM presupuesto WHERE obra_id=? ORDER BY id",
                           (obra_id,))
            editor_tabla("presupuesto", "presupuesto", df, ["partida", "concepto", "monto"],
                         defaults={"obra_id": obra_id})
    with tabs[6]:
        if obra_id == -1:
            st.caption("Selecciona una obra activa.")
        else:
            df = consultar("SELECT id,contratista,concepto,monto_contratado,pagado,avance,estatus "
                           "FROM destajos WHERE obra_id=? ORDER BY id", (obra_id,))
            editor_tabla("destajos", "destajos", df, ["contratista", "concepto",
                         "monto_contratado", "pagado", "avance", "estatus"],
                         defaults={"obra_id": obra_id})
    with tabs[7]:
        if obra_id == -1:
            st.caption("Selecciona una obra activa.")
        else:
            df = consultar("SELECT id,folio,fecha,material,cantidad,unidad,proveedor,"
                           "costo_estimado,estatus FROM requisiciones WHERE obra_id=? ORDER BY id",
                           (obra_id,))
            editor_tabla("requisiciones", "requisiciones", df, ["folio", "fecha", "material",
                         "cantidad", "unidad", "proveedor", "costo_estimado", "estatus"],
                         defaults={"obra_id": obra_id, "solicitante": usuario})
    with tabs[8]:
        if obra_id == -1:
            st.caption("Selecciona una obra activa.")
        else:
            df = consultar("SELECT id,etapa,inicio,fin,estado,avance FROM etapas WHERE obra_id=? ORDER BY id",
                           (obra_id,))
            editor_tabla("etapas", "etapas", df, ["etapa", "inicio", "fin", "estado", "avance"],
                         defaults={"obra_id": obra_id})
    with tabs[9]:
        if obra_id == -1:
            st.caption("Selecciona una obra activa.")
        else:
            df = consultar("SELECT id,fecha,autor,nota FROM bitacora WHERE obra_id=? ORDER BY id",
                           (obra_id,))
            editor_tabla("bitacora", "bitacora", df, ["fecha", "autor", "nota"],
                         defaults={"obra_id": obra_id})
    with tabs[10]:
        if not puede(rol, "usuarios"):
            st.caption("Solo el Administrador puede editar usuarios.")
        else:
            st.caption("Aquí puedes editar datos/rol/obra o borrar usuarios. Para cambiar "
                       "la clave usa «Usuarios → Restablecer la clave».")
            df = consultar("SELECT id,codigo,nombre,rol,telefono,correo,obra_id FROM usuarios ORDER BY id")
            editor_tabla("usuarios", "usuarios", df,
                         ["codigo", "nombre", "rol", "telefono", "correo", "obra_id"])

    # ---- Reiniciar app (borrar todo) ----
    if puede(rol, "admin"):
        st.markdown("---")
        st.markdown("#### ⚠️ Reiniciar aplicación (borrar TODOS los datos)")
        with st.expander("Dejar la app como recién adquirida"):
            st.warning("Esto elimina TODAS las obras, clientes, compras, comprobantes, "
                       "etc. No se puede deshacer.")
            confirma = st.checkbox("Entiendo que se borrará todo y quiero continuar")
            if st.button("🗑️ Borrar todos los datos ahora"):
                if confirma:
                    borrar_todo()
                    st.success("Listo. La app quedó vacía, como nueva."); st.rerun()
                else:
                    st.error("Marca la casilla de confirmación primero.")


# =============================================================================
# 9) REPORTE IMPRIMIBLE
# =============================================================================
def generar_reporte_html(obra_id: int) -> str:
    k = calcular_kpis(obra_id); o = k["obra"]
    obras = obtener_obras().set_index("id")
    cliente = obras.loc[obra_id, "cliente"] if obra_id in obras.index else ""
    etapas = consultar("SELECT etapa,inicio,fin,estado,avance FROM etapas WHERE obra_id=?", (obra_id,))
    dest = consultar("SELECT contratista,concepto,monto_contratado,pagado,avance,estatus "
                     "FROM destajos WHERE obra_id=?", (obra_id,))
    reqs = consultar("SELECT folio,fecha,material,cantidad,unidad,proveedor,costo_estimado,estatus "
                     "FROM requisiciones WHERE obra_id=?", (obra_id,))
    pres = consultar("SELECT partida,concepto,monto FROM presupuesto WHERE obra_id=?", (obra_id,))
    compras = consultar("SELECT categoria, SUM(importe) AS total FROM compras WHERE obra_id=? "
                        "GROUP BY categoria", (obra_id,))
    bit = consultar("SELECT fecha,autor,nota FROM bitacora WHERE obra_id=? ORDER BY fecha DESC",
                    (obra_id,))

    def tabla(df, titulo):
        if df.empty:
            return f"<h3>{titulo}</h3><p>Sin registros.</p>"
        return f"<h3>{titulo}</h3>" + df.to_html(index=False, border=0)

    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">
    <title>Reporte de obra</title><style>
      body{{font-family:'Segoe UI',Arial,sans-serif;color:#2A2A28;margin:40px;}}
      h1{{color:#2F6F6A;border-bottom:3px solid #C9842B;padding-bottom:8px;}}
      h3{{color:#2F6F6A;margin-top:24px;}}
      table{{border-collapse:collapse;width:100%;margin-top:8px;font-size:13px;}}
      th{{background:#2F6F6A;color:#fff;text-align:left;padding:8px;}}
      td{{border-bottom:1px solid #E6E2DA;padding:7px;}}
      .kpis{{display:flex;gap:16px;flex-wrap:wrap;margin:16px 0;}}
      .kpi{{background:#F4F2EE;border:1px solid #E6E2DA;border-radius:10px;padding:14px 18px;min-width:150px;}}
      .kpi b{{display:block;font-size:1.4rem;color:#2F6F6A;}}
      .pie{{margin-top:30px;color:#7A766E;font-size:12px;}}
      @media print {{ button{{display:none;}} }}
    </style></head><body>
    <button onclick="window.print()" style="background:#2F6F6A;color:#fff;border:0;
      padding:10px 20px;border-radius:8px;cursor:pointer;">🖨️ Imprimir / Guardar PDF</button>
    <h1>Reporte de Obra</h1>
    <p><b>Obra:</b> {o['nombre']}<br><b>Cliente:</b> {cliente}<br>
       <b>Ubicación:</b> {o['ubicacion']} &nbsp; <b>Responsable:</b> {o['ingeniero']}<br>
       <b>Fecha del reporte:</b> {HOY.isoformat()}</p>
    <div class="kpis">
      <div class="kpi">Avance general<b>{k['avance']}%</b></div>
      <div class="kpi">Presupuesto<b>{pesos(k['presupuesto'])}</b></div>
      <div class="kpi">Ejercido<b>{pesos(k['ejercido'])} ({k['pct']}%)</b></div>
      <div class="kpi">Días restantes<b>{k['dias']}</b></div>
    </div>
    {tabla(etapas, "Cronograma y avance por etapa")}
    {tabla(pres, "Presupuesto por partida")}
    {tabla(compras, "Compras por categoría")}
    {tabla(dest, "Destajos de contratistas")}
    {tabla(reqs, "Requisiciones de materiales")}
    {tabla(bit, "Bitácora de obra")}
    <p class="pie">Generado por SONA CONSTRUCTORES DEL MAYAB</p>
    </body></html>"""


def vista_reportes(obra_id: int):
    st.subheader("🖨️ Reportes imprimibles")
    if not requiere_obra(obra_id):
        return
    st.write("Genera un reporte completo de la obra. Descárgalo, ábrelo en tu navegador y usa "
             "el botón Imprimir / Guardar PDF (o Ctrl + P).")
    html = generar_reporte_html(obra_id)
    nombre = consultar("SELECT nombre FROM obras WHERE id=?", (obra_id,))["nombre"].iloc[0]
    base = nombre[:25].replace(' ', '_')
    cpdf, chtml = st.columns(2)
    with cpdf:
        if FPDF_OK:
            st.download_button("⬇️ Descargar PDF (un clic)", data=pdf_reporte(obra_id),
                               file_name=f"Reporte_{base}.pdf", mime="application/pdf")
        else:
            st.caption("Para PDF de un clic instala: python -m pip install fpdf2")
    with chtml:
        st.download_button("🖨️ Versión HTML (imprimible)", data=html,
                           file_name=f"Reporte_{base}.html", mime="text/html")
    with st.expander("👁️ Vista previa del reporte"):
        try:
            import streamlit.components.v1 as components
            components.html(html, height=600, scrolling=True)
        except Exception:
            st.info("Descarga el reporte para verlo completo en tu navegador.")

    if FPDF_OK:
        st.markdown("---")
        resumen = (f"Reporte de obra: {nombre}\n"
                   f"{EMPRESA} - {HOY.isoformat()}")
        bloque_enviar_reporte(pdf_reporte(obra_id), f"Reporte de obra - {nombre}",
                              f"Reporte_{base}.pdf", resumen, "rep")


# =============================================================================
# 10) PRINCIPAL
# =============================================================================
def main():
    inyectar_estilos()
    inicializar_bd()
    st.sidebar.markdown("<p class='marca-titulo'>🏗️ Constructor PRO</p>"
                        "<p class='marca-sub'>CRM + Obras + Compras · SONA</p>",
                        unsafe_allow_html=True)
    st.sidebar.markdown("---")

    # --- Inicio de sesión (obligatorio) ---
    if not st.session_state.get("auth"):
        pantalla_login()
        return

    auth = st.session_state.auth
    rol = auth["rol"]
    usuario = auth["nombre"]
    obra_asignada = auth["obra_id"]

    # En la nube: traer datos del Google Sheet una vez por sesión (si está configurado)
    if GSHEETS_OK and _secret("auto_pull") and not st.session_state.get("_pulled"):
        try:
            traer_de_sheets()
        except Exception:
            pass
        st.session_state["_pulled"] = True

    st.sidebar.markdown(f"<span class='chip'>{usuario} · {rol}</span>", unsafe_allow_html=True)
    if st.sidebar.button("Cerrar sesión"):
        st.session_state.auth = None
        st.rerun()
    _cloud_ok = GSHEETS_OK and _sheets_url() and (os.path.exists(GCP_JSON)
                                                  or bool(_secret("gcp_service_account")))
    if _cloud_ok and puede(rol, "editar"):
        if st.sidebar.button("☁️ Guardar en la nube"):
            try:
                sincronizar_a_sheets()
                st.sidebar.success("Datos guardados en Google Sheets.")
            except Exception:
                st.sidebar.error("No se pudo guardar en la nube.")
    st.sidebar.markdown("---")

    disponibles = obras_visibles(rol, obra_asignada)
    if disponibles:
        obras_df = obtener_obras()
        sub = obras_df[obras_df["nombre"].isin(disponibles)]
        ob_labels, ob_map = opciones_clave(sub)
        ob_lbl = st.sidebar.selectbox("🏢 Obra activa", ob_labels)
        obra_nombre = ob_map[ob_lbl]
    else:
        obra_nombre = None
        st.sidebar.info("Sin obras disponibles. " +
                        ("Crea una en «Obras»." if puede(rol, "editar")
                         else "Pide al Administrador que te asigne una."))
    obra_id = obra_id_por_nombre(obra_nombre)

    # Menú según el rol: el Administrador ve todo; los demás (residentes), solo lo permitido
    if puede(rol, "admin"):
        secciones = ["Dashboard", "Usuarios", "CRM", "Obras", "Proveedores", "Contratistas",
                     "Compras", "Presupuesto", "Control Financiero", "Requisiciones", "Destajos",
                     "Avances y Bitácora", "Editar / Borrar", "Reportes", "Respaldo",
                     "Google Sheets"]
    elif rol == "Departamento de Control":
        secciones = ["Dashboard", "Proveedores", "Contratistas", "Compras", "Presupuesto",
                     "Control Financiero", "Requisiciones", "Destajos", "Avances y Bitácora",
                     "Reportes"]
    elif rol == "Ingeniero de Obra":
        secciones = ["Proveedores", "Contratistas", "Compras", "Presupuesto", "Requisiciones",
                     "Destajos", "Avances y Bitácora", "Reportes"]
    else:
        secciones = ["Proveedores", "Contratistas", "Presupuesto", "Requisiciones",
                     "Destajos", "Avances y Bitácora", "Reportes"]
    st.sidebar.markdown("---")
    seccion = st.sidebar.radio("📂 Navegación", secciones)

    st.markdown(f"## {obra_nombre if obra_nombre else 'Sin obra seleccionada'}")
    st.caption(f"Sesión: {usuario} ({rol})  ·  {HOY.strftime('%d/%m/%Y')}")
    st.markdown("---")

    if seccion == "Dashboard":
        vista_dashboard(obra_id)
    elif seccion == "Usuarios":
        vista_usuarios(rol)
    elif seccion == "CRM":
        vista_crm(rol)
    elif seccion == "Obras":
        vista_obras(rol)
    elif seccion == "Proveedores":
        vista_proveedores(rol)
    elif seccion == "Contratistas":
        vista_contratistas(rol)
    elif seccion == "Compras":
        vista_compras(obra_id, rol, usuario)
    elif seccion == "Presupuesto":
        vista_presupuesto(obra_id, rol)
    elif seccion == "Control Financiero":
        vista_control_financiero(obra_id, rol)
    elif seccion == "Requisiciones":
        vista_requisiciones(obra_id, rol, usuario)
    elif seccion == "Destajos":
        vista_destajos(obra_id, rol)
    elif seccion == "Avances y Bitácora":
        vista_avances(obra_id, rol, usuario)
    elif seccion == "Editar / Borrar":
        vista_editar(obra_id, rol, usuario)
    elif seccion == "Reportes":
        vista_reportes(obra_id)
    elif seccion == "Respaldo":
        vista_respaldo(rol)
    elif seccion == "Google Sheets":
        vista_sheets(rol)

    st.sidebar.markdown("---")
    st.sidebar.caption("v5 · Usuarios y claves · Datos en cob_data.db")


if __name__ == "__main__":
    main()
